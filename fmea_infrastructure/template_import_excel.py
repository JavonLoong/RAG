"""Bounded, non-executing XLSX template ingestion."""

# XML is parsed only after ZIP limits, path checks, and declaration checks.
# TRY003 is consistent with the stable ValueError-style importer boundary.
# ruff: noqa: TRY003, S314

from __future__ import annotations

import io
import posixpath
import re
import xml.etree.ElementTree as ET
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import PurePosixPath
from types import MappingProxyType
from uuid import NAMESPACE_URL, uuid5
from zipfile import BadZipFile, ZipFile

from core_domain.fmea.filename_policy import validate_filename
from core_domain.fmea.template_migration import (
    ProposedFieldMapping,
    SourceStructureItem,
    TemplateDraft,
    TemplateDraftStatus,
)


class TemplateImportError(ValueError):
    """Stable, safe error returned by bounded Office package inspection."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        super().__init__(message)


@dataclass(frozen=True, slots=True)
class OfficePackageLimits:
    max_source_bytes: int = 8_000_000
    max_members: int = 256
    max_uncompressed_member_bytes: int = 4_000_000
    max_total_uncompressed_bytes: int = 16_000_000
    max_compression_ratio: float = 200.0
    max_sheets: int = 32
    max_rows: int = 10_000
    max_columns: int = 256
    max_cells: int = 100_000
    max_paragraphs: int = 10_000
    max_tables: int = 256
    max_relationships: int = 512
    max_text_length: int = 4_096
    max_structure_items: int = 4_096

    def __post_init__(self) -> None:
        for name in (
            "max_source_bytes",
            "max_members",
            "max_uncompressed_member_bytes",
            "max_total_uncompressed_bytes",
            "max_sheets",
            "max_rows",
            "max_columns",
            "max_cells",
            "max_paragraphs",
            "max_tables",
            "max_relationships",
            "max_text_length",
            "max_structure_items",
        ):
            value = getattr(self, name)
            if isinstance(value, bool) or not isinstance(value, int) or value < 1:
                raise ValueError(f"{name} must be a positive integer")
        if self.max_compression_ratio <= 1:
            raise ValueError("max_compression_ratio must be greater than one")


@dataclass(frozen=True, slots=True)
class InspectedOfficePackage:
    members: Mapping[str, bytes]


_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_R_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XML_DECLARATION = b"<!DOCTYPE"
_ENTITY_DECLARATION = b"<!ENTITY"
_FORMULA = re.compile(rb"<(?:(?:[A-Za-z_][\w.-]*):)?f(?:\s|>)", re.IGNORECASE)


def _error(code: str, message: str) -> TemplateImportError:
    return TemplateImportError(code, message)


def _text(value: object, field: str, *, limit: int = 256) -> str:
    if not isinstance(value, str) or not value.strip():
        raise _error("FMEA_TEMPLATE_IMPORT_INVALID", f"{field} is invalid")
    normalized = value.strip()
    if len(normalized) > limit:
        raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", f"{field} exceeds the configured limit")
    return normalized


def _clock_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="microseconds").replace("+00:00", "Z")


def _parse_xml(raw: bytes, *, label: str) -> ET.Element:
    if _XML_DECLARATION in raw.upper() or _ENTITY_DECLARATION in raw.upper():
        raise _error("FMEA_TEMPLATE_EXECUTABLE_CONTENT", f"{label} contains unsupported XML declarations")
    try:
        return ET.fromstring(raw)
    except ET.ParseError as exc:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", f"{label} is malformed") from exc


def _unsafe_member_name(name: str) -> bool:
    path = PurePosixPath(name)
    return (
        not name
        or "\\" in name
        or name.startswith("/")
        or any(part in {"", ".", ".."} for part in path.parts)
        or "\x00" in name
    )


def _contains_external_relationship(raw: bytes) -> bool:
    root = _parse_xml(raw, label="relationship part")
    for relationship in root.findall(f"{{{_REL_NAMESPACE}}}Relationship"):
        target_mode = relationship.attrib.get("TargetMode", "").casefold()
        target = relationship.attrib.get("Target", "")
        if target_mode == "external" or re.match(r"(?i)^(?:[a-z][a-z0-9+.-]*:|//|\\\\)", target):
            return True
    return False


def _validate_content_types(raw: bytes, *, kind: str) -> None:
    root = _parse_xml(raw, label="content types")
    content_types = {
        element.attrib.get("ContentType", "").casefold()
        for element in root
        if element.tag.rsplit("}", 1)[-1] == "Override"
    }
    required = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml",
    }.get(kind)
    if required is None or required.casefold() not in content_types:
        raise _error("FMEA_TEMPLATE_CONTENT_TYPE_UNSUPPORTED", "Office content types are unsupported")


def _resolve_part_target(base: str, target: str) -> str:
    if not target or target.startswith("/") or any(part in {"", ".", ".."} for part in target.split("/")):
        raise _error("FMEA_TEMPLATE_PATH_INVALID", "Office relationship target path is invalid")
    resolved = posixpath.normpath(posixpath.join(base, target))
    if _unsafe_member_name(resolved):
        raise _error("FMEA_TEMPLATE_PATH_INVALID", "Office relationship target path is invalid")
    return resolved


def _worksheet_payloads(package: InspectedOfficePackage) -> tuple[tuple[str, bytes], ...]:
    workbook = _parse_xml(package.members["xl/workbook.xml"], label="workbook")
    relationships = package.members.get("xl/_rels/workbook.xml.rels")
    if relationships is None:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook relationships are missing")
    rel_root = _parse_xml(relationships, label="workbook relationships")
    targets = {
        relationship.attrib.get("Id", ""): relationship.attrib.get("Target", "")
        for relationship in rel_root
        if relationship.tag.rsplit("}", 1)[-1] == "Relationship"
    }
    payloads: list[tuple[str, bytes]] = []
    for sheet in workbook.findall(f"{{{_NAMESPACE}}}sheets/{{{_NAMESPACE}}}sheet"):
        name = sheet.attrib.get("name", "")
        relationship_id = sheet.attrib.get(f"{{{_R_NAMESPACE}}}id", "")
        target = targets.get(relationship_id)
        if not name or not target:
            raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook sheet relationship is invalid")
        part_name = _resolve_part_target("xl", target)
        if not part_name.casefold().startswith("xl/worksheets/") or part_name not in package.members:
            raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook worksheet relationship is invalid")
        payloads.append((name, package.members[part_name]))
    return tuple(payloads)


def inspect_office_zip(  # noqa: C901
    raw_bytes: bytes,
    filename: str,
    *,
    kind: str,
    limits: OfficePackageLimits | None = None,
) -> InspectedOfficePackage:
    """Inspect and fully bound the ZIP package before any Office library opens it."""

    active_limits = limits or OfficePackageLimits()
    if type(raw_bytes) is not bytes:
        raise _error("FMEA_TEMPLATE_IMPORT_INVALID", "source bytes are invalid")
    if len(raw_bytes) > active_limits.max_source_bytes:
        raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "source bytes exceed the configured limit")
    try:
        validate_filename(filename, "source_filename", expected_extension=kind)
    except ValueError as exc:
        raise _error("FMEA_TEMPLATE_IMPORT_INVALID", "source filename is invalid") from exc
    try:
        archive = ZipFile(io.BytesIO(raw_bytes), "r")
    except (BadZipFile, OSError, ValueError) as exc:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "Office container is malformed") from exc
    with archive:
        infos = archive.infolist()
        if len(infos) > active_limits.max_members:
            raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "Office container has too many members")
        members: dict[str, bytes] = {}
        total = 0
        for info in infos:
            name = info.filename
            if _unsafe_member_name(name):
                raise _error("FMEA_TEMPLATE_PATH_INVALID", "Office container member path is invalid")
            if name in members:
                raise _error("FMEA_TEMPLATE_DUPLICATE_MEMBER", "Office container has duplicate members")
            if info.flag_bits & 0x1:
                raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "encrypted Office members are unsupported")
            if info.file_size > active_limits.max_uncompressed_member_bytes:
                raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "Office member exceeds the configured limit")
            if info.compress_size and info.file_size > 65_536:
                ratio = info.file_size / info.compress_size
                if ratio > active_limits.max_compression_ratio:
                    raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "Office member compression ratio is unsafe")
            total += info.file_size
            if total > active_limits.max_total_uncompressed_bytes:
                raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "Office container exceeds the configured limit")
            try:
                members[name] = archive.read(info)
            except (BadZipFile, OSError, RuntimeError, ValueError) as exc:
                raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "Office member cannot be read") from exc

    lower_names = {name.casefold() for name in members}
    if any(name.endswith("/vbaproject.bin") or name == "vbaproject.bin" for name in lower_names):
        raise _error("FMEA_TEMPLATE_MACRO_UNSUPPORTED", "macro-enabled Office content is unsupported")
    if any("/embeddings/" in name or name.startswith("embeddings/") for name in lower_names):
        raise _error("FMEA_TEMPLATE_EXECUTABLE_CONTENT", "embedded executable Office content is unsupported")
    for name, payload in members.items():
        if name.casefold().endswith(".rels") and _contains_external_relationship(payload):
            raise _error("FMEA_TEMPLATE_EXTERNAL_CONTENT", "external Office relationships are unsupported")
    content_types = members.get("[Content_Types].xml")
    if content_types is None:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "Office content types are missing")
    _validate_content_types(content_types, kind=kind)
    content_text = content_types.decode("utf-8", errors="replace").casefold()
    if "macroenabled" in content_text or "vbaproject" in content_text:
        raise _error("FMEA_TEMPLATE_MACRO_UNSUPPORTED", "macro-enabled Office content is unsupported")
    if kind == "xlsx" and any("xl/externallinks/" in name for name in lower_names):
        raise _error("FMEA_TEMPLATE_EXTERNAL_CONTENT", "external workbook links are unsupported")
    if kind == "docx" and "word/document.xml" not in members:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "Word document body is missing")
    if kind == "xlsx" and "xl/workbook.xml" not in members:
        raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook body is missing")
    return InspectedOfficePackage(MappingProxyType(members))


_FIELD_ALIASES: dict[str, frozenset[str]] = {
    "item": frozenset({"item", "item id", "item no", "编号", "项目"}),
    "function": frozenset({"function", "功能"}),
    "failure_mode": frozenset({"failure mode", "failure modes", "失效模式"}),
    "causes": frozenset({"cause", "causes", "失效原因", "原因"}),
    "mechanisms": frozenset({"cause", "mechanism", "mechanisms", "cause mechanism", "机理"}),
    "effects": frozenset({"effect", "effects", "失效影响", "影响"}),
    "symptoms": frozenset({"symptom", "symptoms", "现象"}),
    "controls": frozenset({"control", "controls", "现有控制"}),
    "barriers": frozenset({"barrier", "barriers", "屏障"}),
    "recommended_actions": frozenset({"recommended action", "recommended actions", "建议措施"}),
}


def _field_matches(label: str) -> tuple[str, ...]:
    normalized = " ".join(label.casefold().split())
    return tuple(field for field, aliases in _FIELD_ALIASES.items() if normalized in aliases)


def classify_source_fields(
    headers: tuple[tuple[str, str], ...],
) -> tuple[tuple[ProposedFieldMapping, ...], tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    proposed: list[ProposedFieldMapping] = []
    unknown: list[str] = []
    ambiguous: list[str] = []
    identified: list[str] = []
    seen_source_keys: set[str] = set()
    for label, locator in headers:
        matches = _field_matches(label)
        if not matches:
            unknown.append(label)
        elif len(matches) > 1 or label in seen_source_keys:
            ambiguous.append(label)
            for field in matches:
                if field not in identified:
                    identified.append(field)
        else:
            field = matches[0]
            proposed.append(ProposedFieldMapping(source_key=label, target_field=field, source_locator=locator))
            if field not in identified:
                identified.append(field)
        seen_source_keys.add(label)
    return tuple(proposed), tuple(unknown), tuple(ambiguous), tuple(identified)


def _stable_draft_id(workspace_id: str, source_hash: str) -> str:
    return f"draft-{uuid5(NAMESPACE_URL, f'fmea-template:{workspace_id}:{source_hash}')}"


class ExcelTemplateImporter:
    """Read worksheet structure without evaluating formulas, links, or macros."""

    def __init__(
        self,
        *,
        limits: OfficePackageLimits | None = None,
        clock: Callable[[], str] = _clock_now,
        max_uncompressed_member_bytes: int | None = None,
    ) -> None:
        self._limits = limits or OfficePackageLimits(
            max_uncompressed_member_bytes=max_uncompressed_member_bytes
            if max_uncompressed_member_bytes is not None
            else OfficePackageLimits().max_uncompressed_member_bytes
        )
        self._clock = clock

    def parse(self, raw_bytes: bytes, filename: str, *, workspace_id: str) -> TemplateDraft:  # noqa: C901
        workspace = _text(workspace_id, "workspace_id")
        package = inspect_office_zip(raw_bytes, filename, kind="xlsx", limits=self._limits)
        worksheet_payloads = _worksheet_payloads(package)
        if len(worksheet_payloads) > self._limits.max_sheets:
            raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "workbook has too many sheets")
        if not worksheet_payloads:
            raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook has no worksheets")
        for _name, payload in worksheet_payloads:
            if _FORMULA.search(payload):
                raise _error("FMEA_TEMPLATE_FORMULA_UNSUPPORTED", "formula content is unsupported")

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                io.BytesIO(raw_bytes),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except TemplateImportError:
            raise
        except Exception as exc:
            raise _error("FMEA_TEMPLATE_CONTAINER_INVALID", "workbook cannot be parsed") from exc

        structure: list[SourceStructureItem] = []
        headers: list[tuple[str, str]] = []
        total_cell_count = 0
        try:
            if len(workbook.sheetnames) > self._limits.max_sheets:
                raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "workbook has too many sheets")
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                structure.append(SourceStructureItem(kind="sheet", locator=sheet_name))
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                if max_row > self._limits.max_rows or max_column > self._limits.max_columns:
                    raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "worksheet dimensions exceed the configured limit")
                cell_count = 0
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        cell_count += 1
                        total_cell_count += 1
                        if total_cell_count > self._limits.max_cells:
                            raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "worksheet has too many cells")
                        value = cell.value if isinstance(cell.value, str) else str(cell.value)
                        if len(value) > self._limits.max_text_length:
                            raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "cell text exceeds the configured limit")
                        structure.append(
                            SourceStructureItem(kind="cell", locator=f"{sheet_name}!{cell.coordinate}", value=value)
                        )
                        if cell.row == 1 and isinstance(cell.value, str) and cell.value.strip():
                            headers.append((cell.value.strip(), f"{sheet_name}!{cell.coordinate}"))
                worksheet_xml = _parse_xml(worksheet_payloads[sheet_index][1], label="worksheet")
                for merged in worksheet_xml.findall(f"{{{_NAMESPACE}}}mergeCells/{{{_NAMESPACE}}}mergeCell"):
                    reference = merged.attrib.get("ref")
                    if reference:
                        structure.append(SourceStructureItem(kind="merge", locator=f"{sheet_name}!{reference}"))
        finally:
            workbook.close()
        if len(structure) > self._limits.max_structure_items:
            raise _error("FMEA_TEMPLATE_LIMIT_EXCEEDED", "worksheet structure exceeds the configured limit")
        proposed, unknown, ambiguous, identified = classify_source_fields(tuple(headers))
        source_hash = sha256(raw_bytes).hexdigest()
        return TemplateDraft(
            draft_id=_stable_draft_id(workspace, source_hash),
            workspace_id=workspace,
            source_filename=validate_filename(filename, "source_filename", expected_extension="xlsx"),
            source_sha256=source_hash,
            source_type="xlsx",
            structure=tuple(structure),
            proposed_fields=proposed,
            unknown_fields=unknown,
            ambiguous_fields=ambiguous,
            parser_warnings=(),
            status=TemplateDraftStatus.DRAFT,
            created_at=self._clock(),
            identified_fields=identified,
        )


__all__ = [
    "ExcelTemplateImporter",
    "InspectedOfficePackage",
    "OfficePackageLimits",
    "TemplateImportError",
    "classify_source_fields",
    "inspect_office_zip",
]
