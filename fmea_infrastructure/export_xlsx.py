"""Presentation-only XLSX rendering for normalized FMEA snapshots."""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from typing import Final, NoReturn
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

import openpyxl
import orjson
from defusedxml.ElementTree import fromstring as safe_xml_fromstring
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fmea_application.snapshot_contracts import NormalizedFmeaSnapshot

from .export_json import _ensure_snapshot_hash, _snapshot_projection, _validate_export_value

_EXPORT_SCHEMA: Final = "graphrag.fmea.export.v1"
_PREVIEW_MARKER: Final = "DRAFT PREVIEW — NOT PUBLISHED"
_MAX_EXCEL_CELL_TEXT: Final = 32_767
_MAX_COLUMNS: Final = 256
_MAX_WIDTH: Final = 48
_MIN_WIDTH: Final = 12
_WIDTH_SAMPLE_ROWS: Final = 32
_TYPES_COLUMN: Final = "__types__"
_RESERVED_HEADERS: Final = frozenset({"Identity", _TYPES_COLUMN})


class XlsxExportError(ValueError):
    """Stable, public-safe XLSX rendering failure."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        super().__init__(message)


def _error(code: str, message: str) -> XlsxExportError:
    return XlsxExportError(code, message)


def _invalid() -> NoReturn:
    raise ValueError


def _is_xml_char(value: str) -> bool:
    return all(
        codepoint in {0x9, 0xA, 0xD}
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
        for codepoint in map(ord, value)
    )


def _validate_office_value(value: object) -> None:
    if isinstance(value, str):
        if not _is_xml_char(value):
            _invalid()
        return
    if isinstance(value, Mapping):
        for key, item in value.items():
            if not isinstance(key, str) or not _is_xml_char(key):
                _invalid()
            _validate_office_value(item)
        return
    if isinstance(value, Sequence) and not isinstance(value, str | bytes):
        for item in value:
            _validate_office_value(item)


def _json_text(value: object) -> str:
    try:
        return orjson.dumps(value, option=orjson.OPT_SORT_KEYS).decode("utf-8")
    except (TypeError, ValueError, OverflowError):
        _invalid()


def _value_encoding(value: object) -> tuple[str, str]:
    if type(value) is str:
        return value, "str"
    if value is None:
        return "null", "null"
    if type(value) is bool:
        return "true" if value else "false", "bool"
    if type(value) is int:
        return str(value), "int"
    if type(value) is float:
        return _json_text(value), "float"
    if isinstance(value, Mapping | list):
        return _json_text(value), "json"
    _invalid()


def _cell_text(value: str) -> str:
    if len(value) > _MAX_EXCEL_CELL_TEXT:
        _invalid()
    return value


def _set_string_cell(cell, value: str) -> None:
    cell.value = _cell_text(value)
    cell.data_type = "s"
    cell.number_format = "@"
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _identity(record: Mapping[str, object], identity_field: str | None, index: int) -> str:
    if identity_field is not None:
        value = record.get(identity_field)
        if type(value) is not str or not value.strip():
            _invalid()
        return value
    for field_name in ("source_id", "code", "item_id"):
        value = record.get(field_name)
        if type(value) is str and value.strip():
            return value
    return f"item-{index:03d}"


def _record_columns(records: Sequence[Mapping[str, object]]) -> list[str]:
    keys = {key for record in records for key in record}
    if any(type(key) is not str or key in _RESERVED_HEADERS for key in keys):
        _invalid()
    columns = sorted(keys)
    if len(columns) + 2 > _MAX_COLUMNS:
        _invalid()
    return columns


def _style_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(bottom=Side(style="thin", color="D9EAF7"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    end_row = max(1, worksheet.max_row)
    end_column = get_column_letter(max(1, worksheet.max_column))
    worksheet.auto_filter.ref = f"A1:{end_column}{end_row}"
    worksheet.sheet_view.showGridLines = False
    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        sample = [worksheet.cell(row, column_index).value for row in range(1, min(end_row, _WIDTH_SAMPLE_ROWS) + 1)]
        width = min(
            _MAX_WIDTH, max(_MIN_WIDTH, max((len(str(value)) for value in sample if value is not None), default=0) + 2)
        )
        worksheet.column_dimensions[letter].width = width
    if worksheet.max_column >= 1 and worksheet.cell(1, worksheet.max_column).value == _TYPES_COLUMN:
        worksheet.column_dimensions[get_column_letter(worksheet.max_column)].hidden = True


def _append_manifest(worksheet, projection: Mapping[str, object], *, draft_preview: bool) -> None:
    headers = ("Key", "Value", "Type")
    for column, value in enumerate(headers, start=1):
        _set_string_cell(worksheet.cell(1, column), value)
    metadata: tuple[tuple[str, object], ...] = (
        ("schema_version", projection["schema_version"]),
        ("snapshot_schema_version", projection.get("snapshot_schema_version", "graphrag.fmea.normalized-snapshot.v1")),
        ("snapshot_id", projection["snapshot_id"]),
        ("workspace_id", projection["workspace_id"]),
        ("analysis_id", projection["analysis_id"]),
        ("revision_id", projection["revision_id"]),
        ("revision_hash", projection["revision_hash"]),
        ("publication_id", projection["publication_id"]),
        ("manifest_id", projection["manifest_id"]),
        ("row_count", projection["row_count"]),
        ("risk_count", len(projection["risk_records"])),
        ("propagation_present", projection["propagation"] is not None),
        ("evidence_count", len(projection["evidence_summary"])),
        ("decision_count", len(projection["decision_summary"])),
        ("unresolved_count", len(projection["unresolved_items"])),
        ("snapshot_hash", projection["snapshot_hash"]),
        ("created_at", projection["created_at"]),
        ("version_manifest", projection["version_manifest"]),
        ("audit_summary", projection["audit_summary"]),
        ("draft_preview", draft_preview),
        ("draft_marker", _PREVIEW_MARKER if draft_preview else ""),
        ("format", "xlsx"),
        ("media_type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )
    for row_index, (key, value) in enumerate(metadata, start=2):
        encoded, value_type = _value_encoding(value)
        _set_string_cell(worksheet.cell(row_index, 1), key)
        _set_string_cell(worksheet.cell(row_index, 2), encoded)
        _set_string_cell(worksheet.cell(row_index, 3), value_type)


def _append_typed_table(
    worksheet,
    records: Sequence[Mapping[str, object]],
    *,
    identity_field: str | None,
) -> None:
    columns = _record_columns(records)
    headers = ["Identity", *columns, _TYPES_COLUMN]
    for column_index, header in enumerate(headers, start=1):
        _set_string_cell(worksheet.cell(1, column_index), header)
    for row_index, record in enumerate(records, start=2):
        _set_string_cell(worksheet.cell(row_index, 1), _identity(record, identity_field, row_index - 1))
        types: dict[str, str] = {}
        for column_index, key in enumerate(columns, start=2):
            if key not in record:
                _set_string_cell(worksheet.cell(row_index, column_index), "")
                continue
            encoded, value_type = _value_encoding(record[key])
            types[key] = value_type
            _set_string_cell(worksheet.cell(row_index, column_index), encoded)
        _set_string_cell(worksheet.cell(row_index, len(headers)), _json_text(types))


def _validate_package_xml(name: str, raw: bytes) -> None:
    folded_name = name.casefold()
    if not folded_name.endswith((".xml", ".rels")):
        return
    root = safe_xml_fromstring(raw, forbid_dtd=True, forbid_entities=True, forbid_external=True)
    if folded_name.endswith(".rels"):
        for relationship in root.iter("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
            target = relationship.attrib.get("Target", "")
            if relationship.attrib.get("TargetMode", "").casefold() == "external" or target.casefold().startswith((
                "http:",
                "https:",
                "file:",
                "ftp:",
            )):
                _invalid()
    if folded_name.endswith(".xml"):
        for element in root.iter():
            if element.tag.rsplit("}", 1)[-1].casefold() == "f":
                _invalid()


def _validate_package(payload: bytes) -> None:
    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            names = tuple(archive.namelist())
            folded = {name.casefold() for name in names}
            if any(".." in name.split("/") or "\\" in name or name.startswith("/") for name in names):
                _invalid()
            if any(name.startswith("xl/externallinks/") for name in folded):
                _invalid()
            if any(name.endswith((".xlsm", ".xlam", ".bin")) or "vbaproject" in name for name in folded):
                _invalid()
            for name in names:
                _validate_package_xml(name, archive.read(name))
    except (BadZipFile, OSError, ValueError, ParseError):
        _invalid()


class XlsxFmeaExporter:
    """Render a normalized snapshot to an in-memory, presentation-only XLSX."""

    format = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self, draft_preview: bool = False) -> None:
        if type(draft_preview) is not bool:
            raise _error("FMEA_EXPORT_XLSX_INVALID", "draft_preview must be a boolean")
        self._draft_preview = draft_preview

    def render(self, snapshot: NormalizedFmeaSnapshot) -> bytes:
        if type(snapshot) is not NormalizedFmeaSnapshot:
            raise _error("FMEA_EXPORT_SNAPSHOT_INVALID", "snapshot must be a NormalizedFmeaSnapshot")
        try:
            projection = _snapshot_projection(snapshot)
            _validate_export_value(projection)
            _validate_office_value(projection)
            _ensure_snapshot_hash(snapshot)
            workbook = openpyxl.Workbook()
            manifest = workbook.active
            manifest.title = "Manifest"
            for sheet_name in ("FMEA", "Risk", "Propagation", "Evidence", "Decisions", "Unresolved"):
                workbook.create_sheet(sheet_name)
            _append_manifest(manifest, projection, draft_preview=self._draft_preview)
            _append_typed_table(workbook["FMEA"], projection["rows"], identity_field="row_id")
            _append_typed_table(workbook["Risk"], projection["risk_records"], identity_field="assessment_id")
            propagation = () if projection["propagation"] is None else (projection["propagation"],)
            _append_typed_table(workbook["Propagation"], propagation, identity_field=None)
            _append_typed_table(workbook["Evidence"], projection["evidence_summary"], identity_field="pack_id")
            _append_typed_table(workbook["Decisions"], projection["decision_summary"], identity_field="decision_id")
            _append_typed_table(workbook["Unresolved"], projection["unresolved_items"], identity_field=None)
            for worksheet in workbook.worksheets:
                _style_sheet(worksheet)
            buffer = io.BytesIO()
            workbook.save(buffer)
            payload = buffer.getvalue()
            _validate_package(payload)
        except XlsxExportError:
            raise
        except Exception:
            raise _error("FMEA_EXPORT_XLSX_INVALID", "snapshot cannot be rendered as XLSX") from None
        else:
            return payload


__all__ = ["XlsxExportError", "XlsxFmeaExporter"]
