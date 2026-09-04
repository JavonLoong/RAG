"""Independent, fail-closed verification of full FMEA acceptance artifacts.

No runner, fixture builder, or exporter helper is imported here. Inventory
verification alone is deliberately insufficient to accept a product workflow.
"""

# ruff: noqa: RUF001

from __future__ import annotations

import argparse
import io
import json
import os
import re
import stat
from dataclasses import dataclass
from hashlib import sha256
from itertools import pairwise
from pathlib import Path
from uuid import UUID
from zipfile import BadZipFile, ZipFile

import openpyxl
from defusedxml.ElementTree import fromstring as safe_xml_fromstring

LEGACY_SCHEMA_VERSION = "graphrag.fmea.full.acceptance.v1"
SCHEMA_VERSION = "graphrag.fmea.full.acceptance.v2"
BODY_SCHEMA_VERSION = "graphrag.fmea.body.v1"
DEFAULT_OUTPUT_ROOT = Path(__file__).resolve().parents[1] / "observability/reports/fmea-full-acceptance"
_MAX_FILE_BYTES = 32 * 1024 * 1024
_MAX_TOTAL_BYTES = 128 * 1024 * 1024
_MAX_MANIFEST_BYTES = 256 * 1024
_MAX_FILES = 64
_REPARSE = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
_HASH = re.compile(r"^[0-9a-f]{64}$")
_RELATIVE_FILE = re.compile(
    r"^(?:evidence\.json|exports/[a-z0-9][a-z0-9._-]{0,127}\.(?:json|xlsx|docx)|inputs/[a-z0-9][a-z0-9._-]{0,127}\.xlsx)$"
)
_PRIVATE = re.compile(r"(?:\bBearer\s+\S+|(?<![A-Za-z0-9])[A-Za-z]:[\\/]|\\\\|\bfile://)", re.IGNORECASE)
_PRIVATE_KEYS = {"access_token", "api_key", "authorization", "credentials", "password", "private_key", "private_path", "secret", "raw_output", "provider_output", "prompt"}
_TABLES = ("Manifest", "FMEA", "Risk", "Propagation", "Evidence", "Decisions", "Unresolved")
_OFFICE_TABLES = ("正文", "正文详情", *_TABLES)
_WORD = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_LAYOUT_ROW_FIELDS = frozenset(
    {
        "row_id",
        "analysis_id",
        "evidence_pack_id",
        "item_id",
        "function_id",
        "failure_mode",
        "causes",
        "mechanisms",
        "effects",
        "symptoms",
        "controls",
        "barriers",
        "actions",
        "claim_status",
        "review_status",
        "publication_status",
        "record_version",
        "row_hash",
    }
)
_LAYOUT_FIELD_KEY = re.compile(r"[A-Za-z_][A-Za-z0-9_.-]*\Z")
_LAYOUT_VALUE_TYPES = frozenset(
    {"string", "integer", "number", "decimal", "boolean", "object", "array", "null", "json", "string[]"}
)
_P0_FIELDS = (
    "model_approval_count",
    "known_without_evidence_count",
    "confirmed_invalid_score_count",
    "accepted_high_risk_evidence_free_edge_count",
)
_FULL_REQUIRED = (
    "analyses",
    "candidates",
    "scoring_rules",
    "evidence_packs",
    "review_decisions",
    "risk_records",
    "propagation_graphs",
    "revisions",
    "submissions",
    "approvals",
    "publications",
    "snapshots",
    "manifests",
    "exports",
    "template_drafts",
    "template_import_sources",
    "migration_reports",
    "migration_results",
    "lifecycle_events",
    "audits",
    "outbox",
    "replays",
    "steps",
)


class VerificationError(ValueError):
    """Stable error code without private filesystem or provider details."""

    def __init__(self, code: str) -> None:
        self.code = code
        super().__init__(code)


@dataclass(frozen=True)
class VerificationResult:
    passed: bool
    artifact_id: str
    error_code: str | None = None


def _require(condition: bool, code: str) -> None:
    if not condition:
        raise VerificationError(code)


def _pairs(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        _require(key not in result, "FMEA_DUPLICATE_JSON_KEY")
        result[key] = value
    return result


def _nonfinite(_value: str) -> None:
    raise VerificationError("FMEA_NONFINITE_JSON_NUMBER")


def _json(raw: bytes) -> object:
    try:
        value = json.loads(raw.decode("utf-8"), object_pairs_hook=_pairs, parse_constant=_nonfinite)
    except (UnicodeError, ValueError, RecursionError) as error:
        if isinstance(error, VerificationError):
            raise
        raise VerificationError("FMEA_INVALID_JSON") from None
    return value


def _parse(raw: bytes) -> dict[str, object]:
    value = _json(raw)
    _require(isinstance(value, dict), "FMEA_ARTIFACT_SCHEMA_INVALID")
    return value


def _plain(info: os.stat_result, *, directory: bool) -> bool:
    return (
        not stat.S_ISLNK(info.st_mode)
        and not getattr(info, "st_file_attributes", 0) & _REPARSE
        and (stat.S_ISDIR(info.st_mode) if directory else stat.S_ISREG(info.st_mode))
    )


def _directory(path: Path) -> Path:
    candidate = path.absolute()
    current = Path(candidate.anchor)
    for part in candidate.parts[1:]:
        _require(part not in {".", ".."}, "FMEA_ARTIFACT_PATH_INVALID")
        current /= part
        _require(_plain(current.lstat(), directory=True), "FMEA_ARTIFACT_PATH_INVALID")
    return candidate


def _read(path: Path, limit: int) -> bytes:
    before = path.lstat()
    _require(_plain(before, directory=False) and before.st_size <= limit, "FMEA_ARTIFACT_BOUNDS")
    descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0) | getattr(os, "O_BINARY", 0))
    with os.fdopen(descriptor, "rb") as stream:
        opened = os.fstat(stream.fileno())
        _require((opened.st_dev, opened.st_ino) == (before.st_dev, before.st_ino), "FMEA_ARTIFACT_PATH_CHANGED")
        raw = stream.read(limit + 1)
        after = os.fstat(stream.fileno())
    _require(len(raw) <= limit and len(raw) == opened.st_size, "FMEA_ARTIFACT_BOUNDS")
    _require(
        (opened.st_size, opened.st_mtime_ns) == (after.st_size, after.st_mtime_ns),
        "FMEA_ARTIFACT_PATH_CHANGED",
    )
    return raw


def _inventory(root: Path) -> set[str]:
    found: set[str] = set()
    for child in root.iterdir():
        info = child.lstat()
        if child.name in {"exports", "inputs"} and _plain(info, directory=True):
            for artifact in child.iterdir():
                _require(_plain(artifact.lstat(), directory=False), "FMEA_ARTIFACT_PATH_INVALID")
                found.add(f"{child.name}/{artifact.name}")
                _require(len(found) <= _MAX_FILES + 1, "FMEA_ARTIFACT_BOUNDS")
        else:
            _require(_plain(info, directory=False), "FMEA_ARTIFACT_PATH_INVALID")
            found.add(child.name)
        _require(len(found) <= _MAX_FILES + 1, "FMEA_ARTIFACT_BOUNDS")
    return found


def load_bundle(directory: str | Path) -> tuple[dict[str, object], dict[str, bytes]]:
    """Verify bounded byte inventory; this does not certify workflow semantics."""
    try:
        root = _directory(Path(directory))
        manifest = _parse(_read(root / "manifest.json", _MAX_MANIFEST_BYTES))
        _require(manifest.get("schema_version") in {LEGACY_SCHEMA_VERSION, SCHEMA_VERSION}, "FMEA_ARTIFACT_SCHEMA_INVALID")
        entries = manifest.get("files")
        _require(isinstance(entries, dict) and 0 < len(entries) <= _MAX_FILES, "FMEA_ARTIFACT_INVENTORY_INVALID")
        _require("evidence.json" in entries, "FMEA_ARTIFACT_INVENTORY_INVALID")
        _require(all(_RELATIVE_FILE.fullmatch(name) is not None for name in entries), "FMEA_ARTIFACT_PATH_INVALID")
        _require(_inventory(root) == {"manifest.json", *entries}, "FMEA_ARTIFACT_INVENTORY_MISMATCH")
        payloads: dict[str, bytes] = {}
        total = 0
        for name, identity in entries.items():
            _require(isinstance(identity, dict) and set(identity) == {"sha256", "size_bytes"}, "FMEA_ARTIFACT_SCHEMA_INVALID")
            digest, size = identity["sha256"], identity["size_bytes"]
            _require(isinstance(digest, str) and _HASH.fullmatch(digest) is not None, "FMEA_ARTIFACT_HASH_INVALID")
            _require(type(size) is int and 0 <= size <= _MAX_FILE_BYTES, "FMEA_ARTIFACT_BOUNDS")
            total += size
            _require(total <= _MAX_TOTAL_BYTES, "FMEA_ARTIFACT_BOUNDS")
            payload = _read(root / name, size)
            _require(len(payload) == size and sha256(payload).hexdigest() == digest, "FMEA_ARTIFACT_HASH_MISMATCH")
            payloads[name] = payload
    except (OSError, TypeError, KeyError) as error:
        raise VerificationError("FMEA_ARTIFACT_INVALID") from error
    return manifest, payloads


def _privacy(value: object, depth: int = 0) -> None:
    _require(depth <= 32, "FMEA_ARTIFACT_BOUNDS")
    if isinstance(value, str):
        _require(_PRIVATE.search(value) is None, "FMEA_PRIVATE_MARKER")
    elif isinstance(value, dict):
        for key, child in value.items():
            _require(key.casefold().replace("-", "_") not in _PRIVATE_KEYS, "FMEA_PRIVATE_MARKER")
            _privacy(key, depth + 1)
            _privacy(child, depth + 1)
    elif isinstance(value, list):
        for child in value:
            _privacy(child, depth + 1)


def _office_members(payload: bytes) -> dict[str, bytes]:
    members: dict[str, bytes] = {}
    with ZipFile(io.BytesIO(payload)) as archive:
        infos = archive.infolist()
        _require(0 < len(infos) <= 512, "FMEA_OFFICE_BOUNDS")
        _require(len({item.filename.casefold() for item in infos}) == len(infos), "FMEA_OFFICE_DUPLICATE_PART")
        total = 0
        for info in infos:
            name = info.filename
            _require(not name.startswith("/") and "\\" not in name and all(part not in {".", "..", ""} for part in name.split("/")), "FMEA_OFFICE_PATH_INVALID")
            _require(not any(marker in name.casefold() for marker in ("vbaproject", "embeddings/", "activex/")), "FMEA_OFFICE_ACTIVE_CONTENT")
            total += info.file_size
            _require(total <= _MAX_TOTAL_BYTES and info.file_size <= _MAX_FILE_BYTES, "FMEA_OFFICE_BOUNDS")
            with archive.open(info) as stream:
                content = stream.read(info.file_size + 1)
            _require(len(content) == info.file_size, "FMEA_OFFICE_BOUNDS")
            members[name] = content
            if name.casefold().endswith((".xml", ".rels")):
                root = safe_xml_fromstring(content, forbid_dtd=True, forbid_entities=True, forbid_external=True)
                for node in root.iter():
                    tag = node.tag.rsplit("}", 1)[-1].casefold()
                    _require(tag not in {"altchunk", "fldsimple", "fldchar", "instrtext", "f", *_PRIVATE_KEYS}, "FMEA_OFFICE_ACTIVE_CONTENT")
                    _privacy(node.text or "")
                    _privacy(node.tail or "")
                    for value in node.attrib.values():
                        _privacy(value)
                    if tag == "relationship":
                        target = node.attrib.get("Target", "").lower()
                        _require(node.attrib.get("TargetMode", "").lower() != "external" and not re.match(r"^[a-z]+:", target), "FMEA_OFFICE_EXTERNAL_REFERENCE")
    return members


def _decode_cell(value: object, kind: str) -> object:
    if kind == "str":
        return "" if value is None else str(value)
    if kind == "null":
        _require(value in {None, "", "null"}, "FMEA_OFFICE_VALUE_INVALID")
        return None
    _require(kind in {"bool", "int", "float", "json"}, "FMEA_OFFICE_TYPE_INVALID")
    decoded = _json(str(value).encode("utf-8"))
    if kind != "json":
        _require(type(decoded) is {"bool": bool, "int": int, "float": float}[kind], "FMEA_OFFICE_TYPE_INVALID")
    return decoded


def _typed_rows(rows: list[list[object]], identity_field: str | None) -> list[dict[str, object]]:
    _require(bool(rows), "FMEA_OFFICE_TABLE_INVALID")
    headers = rows[0]
    _require(len(headers) == len(set(headers)) and len(headers) >= 2 and headers[0] == "Identity" and headers[-1] == "__types__", "FMEA_OFFICE_TABLE_INVALID")
    result = []
    identities = set()
    for row in rows[1:]:
        _require(len(row) == len(headers), "FMEA_OFFICE_TABLE_INVALID")
        kinds = _parse(str(row[-1]).encode("utf-8"))
        _require(set(kinds) <= set(headers[1:-1]), "FMEA_OFFICE_TABLE_INVALID")
        item = {}
        for name, value in zip(headers[1:-1], row[1:-1], strict=True):
            if name in kinds:
                item[name] = _decode_cell(value, kinds[name])
            else:
                _require(value in {None, ""}, "FMEA_OFFICE_UNDECLARED_VALUE")
        identity = item.get(identity_field) if identity_field else next((item[key] for key in ("source_id", "code", "item_id") if isinstance(item.get(key), str) and item[key].strip()), f"item-{len(result) + 1:03d}")
        _require(isinstance(identity, str) and bool(identity.strip()) and row[0] == identity, "FMEA_OFFICE_IDENTITY_INVALID")
        if identity_field:
            _require(identity not in identities, "FMEA_OFFICE_DUPLICATE_IDENTITY")
            identities.add(identity)
        result.append(item)
    return result


def _word_cell(cell) -> str:
    paragraphs = []
    for paragraph in cell.findall(f"{_WORD}p"):
        parts = []
        for node in paragraph.iter():
            if node.tag == f"{_WORD}t":
                parts.append(node.text or "")
            elif node.tag in {f"{_WORD}br", f"{_WORD}cr"}:
                parts.append("\n")
            elif node.tag == f"{_WORD}tab":
                parts.append("\t")
        paragraphs.append("".join(parts))
    return "\n".join(paragraphs)


def _word_paragraph(node) -> str:
    parts = []
    for child in node.iter():
        if child.tag == f"{_WORD}t":
            parts.append(child.text or "")
        elif child.tag in {f"{_WORD}br", f"{_WORD}cr"}:
            parts.append("\n")
        elif child.tag == f"{_WORD}tab":
            parts.append("\t")
    return "".join(parts)


def _office_projection(tables: list[list[list[object]]]) -> dict[str, object]:
    _require(len(tables) == len(_TABLES) and tables[0][0] == ["Key", "Value", "Type"], "FMEA_OFFICE_TABLE_INVALID")
    metadata = {}
    for row in tables[0][1:]:
        _require(len(row) == 3 and row[0] not in metadata, "FMEA_OFFICE_MANIFEST_INVALID")
        metadata[row[0]] = _decode_cell(row[1], row[2])
    collections = [_typed_rows(rows, identity) for rows, identity in zip(tables[1:], ("row_id", "assessment_id", None, "pack_id", "decision_id", None), strict=True)]
    rows, risk, propagation, evidence, decisions, unresolved = collections
    _require(len(propagation) <= 1, "FMEA_OFFICE_TABLE_INVALID")
    for count, actual in (("row_count", len(rows)), ("risk_count", len(risk)), ("evidence_count", len(evidence)), ("decision_count", len(decisions)), ("unresolved_count", len(unresolved))):
        _require(metadata.get(count) == actual, "FMEA_OFFICE_COUNT_MISMATCH")
    _require(metadata.get("propagation_present") is bool(propagation), "FMEA_OFFICE_COUNT_MISMATCH")
    for key in ("risk_count", "evidence_count", "decision_count", "unresolved_count", "propagation_present"):
        metadata.pop(key)
    return {**metadata, "rows": rows, "risk_records": risk, "propagation": propagation[0] if propagation else None, "evidence_summary": evidence, "decision_summary": decisions, "unresolved_items": unresolved}


def parse_export(
    payload: bytes,
    format_name: str,
    *,
    contract_version: str = SCHEMA_VERSION,
) -> dict[str, object]:
    """Read actual Office cells using a parser independent from the exporters."""
    try:
        _require(contract_version in {LEGACY_SCHEMA_VERSION, SCHEMA_VERSION}, "FMEA_ARTIFACT_SCHEMA_INVALID")
        if format_name == "json":
            result = _parse(payload)
        else:
            _require(format_name in {"xlsx", "docx"}, "FMEA_EXPORT_FORMAT_INVALID")
            members = _office_members(payload)
            if format_name == "xlsx":
                workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
                try:
                    names = tuple(workbook.sheetnames)
                    allowed_names = (
                        {_TABLES, _OFFICE_TABLES}
                        if contract_version == LEGACY_SCHEMA_VERSION
                        else {_OFFICE_TABLES}
                    )
                    _require(names in allowed_names, "FMEA_OFFICE_TABLE_INVALID")
                    tables = [[list(row) for row in workbook[name].iter_rows(values_only=True)] for name in _TABLES]
                finally:
                    workbook.close()
            else:
                root = safe_xml_fromstring(members["word/document.xml"], forbid_dtd=True, forbid_entities=True, forbid_external=True)
                canonical_tables = {}
                marker = None
                body = root.find(f"{_WORD}body")
                _require(body is not None, "FMEA_OFFICE_TABLE_INVALID")
                for child in body:
                    if child.tag == f"{_WORD}p":
                        text = _word_paragraph(child).strip()
                        if text.startswith("Canonical table: "):
                            marker = text.removeprefix("Canonical table: ")
                    elif child.tag == f"{_WORD}tbl" and marker is not None:
                        _require(marker in _TABLES and marker not in canonical_tables, "FMEA_OFFICE_TABLE_INVALID")
                        canonical_tables[marker] = [
                            [_word_cell(cell) for cell in row.findall(f"{_WORD}tc")]
                            for row in child.findall(f"{_WORD}tr")
                        ]
                        marker = None
                if set(canonical_tables) == set(_TABLES):
                    tables = [canonical_tables[name] for name in _TABLES]
                else:
                    _require(contract_version == LEGACY_SCHEMA_VERSION, "FMEA_OFFICE_TABLE_INVALID")
                    tables = [
                        [
                            [_word_cell(cell) for cell in row.findall(f"{_WORD}tc")]
                            for row in table.findall(f"{_WORD}tr")
                        ]
                        for table in body.findall(f"{_WORD}tbl")
                    ]
                    _require(len(tables) == len(_TABLES), "FMEA_OFFICE_TABLE_INVALID")
            result = _office_projection(tables)
        _require(result.get("format") == format_name, "FMEA_EXPORT_FORMAT_INVALID")
        _privacy(result)
    except VerificationError:
        raise
    except (BadZipFile, KeyError, ValueError, TypeError, OSError, IndexError) as error:
        raise VerificationError("FMEA_EXPORT_CONTENT_INVALID") from error
    return result


def _unsupported_known_claims(case: dict[str, object], packs: dict[str, set[str]]) -> int:
    seen_claims = set()
    rows = list(case.get("candidates", []))
    for row in rows:
        valid_ids = packs.get(row.get("evidence_pack_id"), set())
        claims = list(row.get("field_claims", []))
        if row.get("claim_status") == "known":
            bindings = dict(row.get("field_evidence", []))
            declared = {claim.get("field_key") for claim in claims}
            claims.extend({"field_key": name, "claim_status": "known", "evidence_ids": bindings.get(name, [])} for name in ("failure_mode", "causes", "mechanisms", "effects", "symptoms", "controls", "barriers", "actions") if row.get(name) and name not in declared)
        for claim in claims:
            key = (row.get("row_id"), row.get("record_version"), claim.get("field_key"))
            evidence_ids = claim.get("evidence_ids", [])
            if claim.get("claim_status") == "known" and (not evidence_ids or not set(evidence_ids) <= valid_ids) and key not in seen_claims:
                seen_claims.add(key)
    return len(seen_claims)


def _invalid_confirmed_score(record: dict[str, object], rules: dict, packs: dict) -> bool:
    if record.get("status") != "confirmed":
        return False
    dimensions = record.get("dimensions", [])
    rule = rules.get((record.get("rule_pack_id"), record.get("rule_pack_version")), {})
    derived = record.get("derived") or {}
    valid_ids = packs.get(record.get("evidence_pack_id"), set())
    valid = len(dimensions) == 3 and {dimension.get("name") for dimension in dimensions} == {"severity", "occurrence", "detection"} and bool(rule)
    for dimension in dimensions:
        value = dimension.get("value")
        evidence_ids = dimension.get("evidence_ids", [])
        valid = valid and type(value) is int and rule["score_min"] <= value <= rule["score_max"] and bool(evidence_ids) and set(evidence_ids) <= valid_ids
    if valid:
        scores = {dimension["name"]: dimension["value"] for dimension in dimensions}
        valid = derived.get("decision_severity") == scores["severity"] and derived.get("occurrence") == scores["occurrence"] and derived.get("detection") == scores["detection"] and derived.get("rpn") == scores["severity"] * scores["occurrence"] * scores["detection"]
    return not valid


def count_p0_violations(case: dict[str, object]) -> dict[str, int]:
    """Recompute diagnostic violation counts; zero is not proof of completeness."""
    packs = {pack["pack_id"]: {ref["evidence_id"] for ref in pack["refs"]} for pack in case.get("evidence_packs", [])}
    rules = {(rule["rule_pack_id"], rule["version"]): rule for rule in case.get("scoring_rules", [])}
    counts = {
        "model_approval_count": sum(audit.get("actor_type") == "model" and audit.get("command") in {"fmea.approval.decide", "fmea.publication.publish"} for audit in case.get("audits", [])),
        "known_without_evidence_count": _unsupported_known_claims(case, packs),
        "confirmed_invalid_score_count": sum(_invalid_confirmed_score(record, rules, packs) for record in case.get("risk_records", [])),
        "accepted_high_risk_evidence_free_edge_count": 0,
    }
    for graph in case.get("propagation_graphs", []):
        for edge in graph.get("edges", []):
            ids = edge.get("evidence_ids", [])
            valid_ids = packs.get(edge.get("evidence_pack_id"), set())
            if edge.get("risk_priority") in {"high", "critical"} and edge.get("review_status") == "accepted" and (not ids or not set(ids) <= valid_ids):
                counts["accepted_high_risk_evidence_free_edge_count"] += 1
    return counts


def _hash_json(value: object, *, ascii_only: bool = False) -> str:
    return sha256(json.dumps(value, ensure_ascii=ascii_only, sort_keys=True, separators=(",", ":"), allow_nan=False).encode("utf-8")).hexdigest()


def _digest(value: object) -> str:
    _require(isinstance(value, str), "FMEA_HASH_INVALID")
    result = value.removeprefix("sha256:")
    _require(_HASH.fullmatch(result) is not None, "FMEA_HASH_INVALID")
    return result


def _required_list(case: dict[str, object], name: str) -> list[dict[str, object]]:
    value = case.get(name)
    _require(isinstance(value, list) and bool(value), "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
    _require(all(isinstance(item, dict) for item in value), "FMEA_ARTIFACT_SCHEMA_INVALID")
    return value


def _text(value: object, code: str = "FMEA_ARTIFACT_SCHEMA_INVALID") -> str:
    _require(isinstance(value, str) and bool(value.strip()), code)
    return value


def _version(value: object) -> None:
    _require(type(value) is int and value >= 1, "FMEA_VERSION_INVALID")


def _scoped(records: list[dict[str, object]], workspace_id: str, analysis_id: str) -> None:
    for record in records:
        for field, expected in (("workspace_id", workspace_id), ("analysis_id", analysis_id)):
            if field in record:
                _require(record[field] == expected, "FMEA_SCOPE_BINDING_INVALID")


def _indexed(records: list[dict[str, object]], identity: str, code: str = "FMEA_DUPLICATE_RECORD") -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for record in records:
        key = _text(record.get(identity), "FMEA_ARTIFACT_ID_INVALID")
        _require(key not in result, code)
        result[key] = record
    return result


def _composite_index(records: list[dict[str, object]], fields: tuple[str, ...]) -> set[tuple[object, ...]]:
    result: set[tuple[object, ...]] = set()
    for record in records:
        key = tuple(record.get(field) for field in fields)
        _require(all(item is not None for item in key), "FMEA_ARTIFACT_SCHEMA_INVALID")
        _require(key not in result, "FMEA_DUPLICATE_RECORD")
        result.add(key)
    return result


def _valid_uuid(value: object) -> bool:
    if not isinstance(value, str):
        return False
    try:
        parsed = UUID(value)
    except (ValueError, AttributeError, TypeError):
        return False
    return parsed.version == 4 and str(parsed) == value.lower()


def resolve_latest_directory(output_root: str | Path = DEFAULT_OUTPUT_ROOT) -> Path:
    """Resolve an atomic latest pointer without permitting path traversal."""
    try:
        root = _directory(Path(output_root))
        pointer = _parse(_read(root / "latest.json", 4096))
        _require(set(pointer) == {"artifact_id"}, "FMEA_LATEST_POINTER_INVALID")
        artifact_id = pointer.get("artifact_id")
        _require(_valid_uuid(artifact_id), "FMEA_LATEST_POINTER_INVALID")
        target = root / artifact_id
        _require(target.parent == root and target.name == artifact_id, "FMEA_LATEST_POINTER_INVALID")
        return _directory(target)
    except VerificationError:
        raise
    except (OSError, TypeError, ValueError):
        raise VerificationError("FMEA_LATEST_POINTER_INVALID") from None


def _replay_semantics_equal(first: dict[str, object], replayed: dict[str, object]) -> bool:
    """Compare native result DTOs, whose only retry delta is ``replayed``."""
    left = dict(first)
    right = dict(replayed)
    if "replayed" not in left and "replayed" not in right:
        return left == right
    _require("replayed" in left and "replayed" in right, "FMEA_REPLAY_SCHEMA_INVALID")
    left_flag = left.pop("replayed", None)
    right_flag = right.pop("replayed", None)
    _require(type(left_flag) is bool and type(right_flag) is bool, "FMEA_REPLAY_SCHEMA_INVALID")
    return left == right


def _validate_evidence_bindings(
    case: dict[str, object], packs: dict[str, dict[str, object]], workspace_id: str, analysis_id: str
) -> set[str]:
    selected = case.get("evidence_selection")
    _require(isinstance(selected, dict), "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(selected.get("mode") == "offline_source_fixture", "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(selected.get("requested_profile") == "rag_only", "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(selected.get("resolved_profile") == "rag_only", "FMEA_EVIDENCE_SELECTION_INVALID")
    selected_ids = selected.get("selected_evidence_ids")
    _require(isinstance(selected_ids, list) and bool(selected_ids), "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(all(isinstance(item, str) and item for item in selected_ids), "FMEA_EVIDENCE_SELECTION_INVALID")
    pack_id = _text(selected.get("pack_id"), "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(pack_id in packs, "FMEA_EVIDENCE_SELECTION_INVALID")
    pack = packs[pack_id]
    _require(selected.get("pack_hash") == pack.get("pack_hash"), "FMEA_EVIDENCE_SELECTION_INVALID")
    ref_ids = {ref["evidence_id"] for ref in pack["refs"]}
    _require(set(selected_ids) <= ref_ids, "FMEA_EVIDENCE_SELECTION_INVALID")
    _require(set(selected_ids) == ref_ids, "FMEA_EVIDENCE_SELECTION_INVALID")
    for pack in packs.values():
        _text(pack.get("pack_id"), "FMEA_ARTIFACT_ID_INVALID")
        _digest(pack.get("pack_hash"))
        _scoped([pack], workspace_id, analysis_id)
        refs = pack.get("refs")
        _require(isinstance(refs, list) and bool(refs), "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
        _indexed(refs, "evidence_id", "FMEA_DUPLICATE_EVIDENCE")
        for ref in refs:
            _text(ref.get("evidence_id"), "FMEA_ARTIFACT_ID_INVALID")
            _digest(ref.get("evidence_hash"))
            if "content_hash" in ref:
                _digest(ref["content_hash"])
            _text(ref.get("locator"))
            _scoped([ref], workspace_id, analysis_id)
    return ref_ids


def _validate_review_and_risk(
    case: dict[str, object],
    packs: dict[str, dict[str, object]],
    audits: dict[str, dict[str, object]],
    outbox: dict[str, dict[str, object]],
    workspace_id: str,
    analysis_id: str,
) -> None:
    candidates = _indexed(case["candidates"], "row_id")
    # The native candidate query already returns the persisted current FmeaRow
    # v2 records.  Do not invent a second ``currentacceptedrows`` projection:
    # every review decision must bind independently to this native row.
    current = candidates
    for row in current.values():
        _require(row.get("review_status") == "accepted", "FMEA_REVIEW_BINDING_INVALID")
        _require(row.get("record_version") == 2, "FMEA_VERSION_INVALID")
        _require(row.get("evidence_pack_id") in packs, "FMEA_EVIDENCE_BINDING_INVALID")
    _scoped(list(candidates.values()) + list(current.values()), workspace_id, analysis_id)

    decisions = _indexed(case["review_decisions"], "decision_id")
    _require(bool(decisions), "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
    for decision in decisions.values():
        row = decision.get("row") if isinstance(decision.get("row"), dict) else decision
        _require(isinstance(row, dict) and row.get("row_id") in current, "FMEA_REVIEW_BINDING_INVALID")
        _require(row.get("review_status") == "accepted", "FMEA_REVIEW_BINDING_INVALID")
        _require(row == current[row["row_id"]], "FMEA_REVIEW_BINDING_INVALID")
        for event_key in ("audit_event_id", "outbox_event_id"):
            if event_key in decision:
                _require(decision[event_key] in audits or decision[event_key] in outbox, "FMEA_AUDIT_BINDING_INVALID")

    risks = case["risk_records"]
    _composite_index(risks, ("assessment_id", "record_version"))
    risk_ids = {record.get("assessment_id") for record in risks}
    _require(len(risk_ids) == 1, "FMEA_RISK_BINDING_INVALID")
    _require({record.get("status") for record in risks} >= {"proposed", "confirmed"}, "FMEA_RISK_BINDING_INVALID")
    rules = {(rule.get("rule_pack_id"), rule.get("version")): rule for rule in case["scoring_rules"]}
    _require(len(rules) == len(case["scoring_rules"]), "FMEA_DUPLICATE_RECORD")
    for record in risks:
        _version(record.get("record_version"))
        _require(record.get("row_id") in current, "FMEA_RISK_BINDING_INVALID")
        _require(record.get("source_record_version") == current[record["row_id"]].get("record_version"), "FMEA_RISK_BINDING_INVALID")
        _require(record.get("evidence_pack_id") in packs, "FMEA_EVIDENCE_BINDING_INVALID")
        _require((record.get("rule_pack_id"), record.get("rule_pack_version")) in rules, "FMEA_RISK_BINDING_INVALID")
        for dimension in record.get("dimensions", []):
            _require(isinstance(dimension, dict), "FMEA_RISK_BINDING_INVALID")
            ids = dimension.get("evidence_ids")
            _require(isinstance(ids, list) and set(ids) <= {ref["evidence_id"] for ref in packs[record["evidence_pack_id"]]["refs"]}, "FMEA_EVIDENCE_BINDING_INVALID")
    confirmed = next(record for record in risks if record.get("status") == "confirmed")
    _text(confirmed.get("confirmer_actor_id"), "FMEA_AUTHORITY_BINDING_INVALID")
    _require(
        any(
            event.get("command") == "fmea.risk.confirm"
            and event.get("actor_type") == "human"
            and event.get("actor_id") == confirmed.get("confirmer_actor_id")
            and event.get("row_id") == confirmed.get("row_id")
            for event in audits.values()
        ),
        "FMEA_AUTHORITY_BINDING_INVALID",
    )
    _require(
        any(event.get("event_type") == "risk.confirmed" and event.get("aggregate_id") == confirmed.get("assessment_id") for event in outbox.values()),
        "FMEA_OUTBOX_BINDING_INVALID",
    )


def _validate_graph_and_revision_links(
    case: dict[str, object],
    packs: dict[str, dict[str, object]],
    revisions: dict[str, dict[str, object]],
    workspace_id: str,
    analysis_id: str,
) -> tuple[dict[str, object], dict[str, object], dict[str, object]]:
    graphs = _indexed(case["propagation_graphs"], "graph_revision_id")
    _require(len(graphs) >= 2, "FMEA_PROPAGATION_BINDING_INVALID")
    proposed = [graph for graph in graphs.values() if graph.get("status") == "proposed"]
    confirmed = [graph for graph in graphs.values() if graph.get("status") == "confirmed"]
    _require(len(proposed) == 1 and len(confirmed) == 1, "FMEA_PROPAGATION_BINDING_INVALID")
    proposed_graph, confirmed_graph = proposed[0], confirmed[0]
    _require(confirmed_graph.get("parent_graph_revision_id") == proposed_graph.get("graph_revision_id"), "FMEA_PROPAGATION_BINDING_INVALID")
    _scoped([proposed_graph, confirmed_graph], workspace_id, analysis_id)
    pack_ids = set(packs)
    for graph in (proposed_graph, confirmed_graph):
        _require(isinstance(graph.get("edges"), list) and bool(graph["edges"]), "FMEA_PROPAGATION_BINDING_INVALID")
        for edge in graph["edges"]:
            _require(isinstance(edge, dict), "FMEA_PROPAGATION_BINDING_INVALID")
            _require(edge.get("evidence_pack_id") in pack_ids, "FMEA_EVIDENCE_BINDING_INVALID")
            evidence_ids = edge.get("evidence_ids")
            _require(isinstance(evidence_ids, list) and bool(evidence_ids), "FMEA_EVIDENCE_BINDING_INVALID")
            valid_ids = {ref["evidence_id"] for ref in packs[edge["evidence_pack_id"]]["refs"]}
            _require(set(evidence_ids) <= valid_ids, "FMEA_EVIDENCE_BINDING_INVALID")
    graph_id = confirmed_graph.get("graph_revision_id")
    matching = [revision for revision in revisions.values() if revision.get("propagation_graph_revision_id") == graph_id]
    _require(matching, "FMEA_PROPAGATION_BINDING_INVALID")
    return proposed_graph, confirmed_graph, matching[0]


def _validate_propagation_receipts(case: dict[str, object]) -> None:
    """Bind native source-row lineage and propagation replay state snapshots."""
    candidates = _indexed(case["candidates"], "row_id")
    bindings = case.get("source_row_bindings")
    _require(isinstance(bindings, list) and bool(bindings), "FMEA_PROPAGATION_BINDING_INVALID")
    binding_index = _indexed(bindings, "row_id", "FMEA_DUPLICATE_RECORD")
    _require(set(binding_index) == set(candidates), "FMEA_PROPAGATION_BINDING_INVALID")
    for row_id, binding in binding_index.items():
        row = candidates[row_id]
        _require(binding.get("record_version") == row.get("record_version"), "FMEA_PROPAGATION_BINDING_INVALID")
        row_hash = _digest(binding.get("row_hash"))
        _require(row_hash == _hash_json(row, ascii_only=False), "FMEA_PROPAGATION_BINDING_INVALID")
        _require(_digest(binding.get("persisted_row_hash_after")) == row_hash, "FMEA_PROPAGATION_BINDING_INVALID")

    lineage = case.get("source_row_lineage")
    _require(isinstance(lineage, list) and bool(lineage), "FMEA_PROPAGATION_LINEAGE_INVALID")
    graphs = _indexed(case["propagation_graphs"], "graph_revision_id", "FMEA_DUPLICATE_RECORD")
    runs = _indexed(case.get("propagation_runs", []), "run_id", "FMEA_DUPLICATE_RECORD")
    lineage_index = _indexed(lineage, "graph_revision_id", "FMEA_DUPLICATE_RECORD")
    _require(set(lineage_index) == set(graphs), "FMEA_PROPAGATION_LINEAGE_INVALID")
    for graph_id, item in lineage_index.items():
        source_row_id = _text(item.get("source_row_id"), "FMEA_PROPAGATION_LINEAGE_INVALID")
        _require(source_row_id in candidates, "FMEA_PROPAGATION_LINEAGE_INVALID")
        _require(item.get("record_version") == candidates[source_row_id].get("record_version"), "FMEA_PROPAGATION_LINEAGE_INVALID")
        _require(_text(item.get("run_id"), "FMEA_PROPAGATION_LINEAGE_INVALID") in runs, "FMEA_PROPAGATION_LINEAGE_INVALID")
        canonical_hash = _digest(item.get("canonical_row_hash"))
        _require(canonical_hash == _hash_json(candidates[source_row_id], ascii_only=False), "FMEA_PROPAGATION_LINEAGE_INVALID")
        _require(graphs[graph_id].get("status") in {"proposed", "confirmed"}, "FMEA_PROPAGATION_LINEAGE_INVALID")

    for replay in case["replays"]:
        if not str(replay.get("command", "")).startswith("fmea.propagation."):
            continue
        before = replay.get("state_hash_before")
        after = replay.get("state_hash_after")
        _require(before is not None and after is not None, "FMEA_REPLAY_SCHEMA_INVALID")
        _require(_digest(before) == _digest(after), "FMEA_REPLAY_STATE_HASH_MISMATCH")


def _validate_revisions(
    case: dict[str, object], workspace_id: str, analysis_id: str
) -> tuple[dict[str, dict[str, object]], dict[str, object], dict[str, object], dict[str, object]]:
    revisions = _indexed(case["revisions"], "revision_id")
    _require(len(revisions) == 3, "FMEA_REVISION_BINDING_INVALID")
    roots = [revision for revision in revisions.values() if revision.get("parent_revision_id") is None]
    _require(len(roots) == 1, "FMEA_REVISION_BINDING_INVALID")
    root = roots[0]
    children = [revision for revision in revisions.values() if revision.get("parent_revision_id") == root.get("revision_id")]
    _require(len(children) == 2, "FMEA_REVISION_BINDING_INVALID")
    _scoped(list(revisions.values()), workspace_id, analysis_id)
    for revision in revisions.values():
        _digest(revision.get("revision_hash"))
        for field in ("row_versions", "risk_versions", "template_identities", "scoring_rule_identities"):
            _require(isinstance(revision.get(field), list), "FMEA_REVISION_BINDING_INVALID")
        if revision is not root:
            _require(revision.get("parent_revision_hash") == root.get("revision_hash"), "FMEA_REVISION_BINDING_INVALID")
    migration_results = case["migration_results"]
    migration_child_ids = {result.get("child_revision_id") for result in migration_results}
    _require(len(migration_child_ids) == 1 and next(iter(migration_child_ids)) in revisions, "FMEA_MIGRATION_BINDING_INVALID")
    migration_child = revisions[next(iter(migration_child_ids))]
    _require(migration_child in children, "FMEA_MIGRATION_BINDING_INVALID")
    governance_children = [child for child in children if child is not migration_child]
    return revisions, root, governance_children[0], migration_child


def _one_event(case: dict, event_type: str, aggregate_id: str) -> dict:
    matches = [event for event in case["outbox"] if event.get("event_type") == event_type and event.get("aggregate_id") == aggregate_id]
    _require(len(matches) == 1, "FMEA_OUTBOX_BINDING_INVALID")
    event = matches[0]
    _require(event.get("workspace_id") == case["revisions"][0]["workspace_id"], "FMEA_OUTBOX_BINDING_INVALID")
    return event


def _validate_legacy_versioned_content(case: dict, revisions: dict, migration_child: dict) -> None:
    """Join immutable identity projections back to their actual native DTOs."""
    rows = sorted(case["candidates"], key=lambda row: row["row_id"])
    risks = sorted((record for record in case["risk_records"] if record["status"] == "confirmed"), key=lambda record: record["assessment_id"])
    row_versions = [[row["row_id"], row["record_version"], _hash_json(row)] for row in rows]
    risk_versions = [[risk["assessment_id"], risk["record_version"], _hash_json(risk)] for risk in risks]
    graph = next(item for item in case["propagation_graphs"] if item["status"] == "confirmed")
    for revision in revisions.values():
        _require(sorted(revision["row_versions"]) == row_versions, "FMEA_REVISION_CONTENT_MISMATCH")
        if revision["revision_id"] == migration_child["revision_id"]:
            _require(revision["risk_versions"] == [] and revision.get("propagation_graph_revision_id") is None and revision.get("propagation_graph_hash") is None, "FMEA_MIGRATION_BINDING_INVALID")
        else:
            _require(sorted(revision["risk_versions"]) == risk_versions, "FMEA_REVISION_CONTENT_MISMATCH")
            _require(revision.get("propagation_graph_revision_id") == graph["graph_revision_id"] and _digest(revision.get("propagation_graph_hash")) == _hash_json(graph), "FMEA_PROPAGATION_CONTENT_MISMATCH")
    for snapshot in case["snapshots"]:
        revision = revisions.get(snapshot.get("revision_id"))
        _require(revision is not None, "FMEA_SNAPSHOT_CONTENT_MISMATCH")
        expected_rows = [{"row_id": identity, "record_version": version, "row_hash": digest} for identity, version, digest in row_versions]
        expected_risks = [{"assessment_id": identity, "record_version": version, "assessment_hash": digest} for identity, version, digest in risk_versions]
        _require(snapshot.get("rows") == expected_rows and snapshot.get("risk_records") == expected_risks, "FMEA_SNAPSHOT_CONTENT_MISMATCH")
        _require(snapshot.get("propagation") == {"graph_revision_id": graph["graph_revision_id"], "graph_hash": _hash_json(graph)}, "FMEA_SNAPSHOT_CONTENT_MISMATCH")
    for risk in risks:
        event = _one_event(case, "risk.confirmed", risk["assessment_id"])
        _require(event.get("payload", {}).get("assessment") == risk, "FMEA_RISK_CONTENT_MISMATCH")
        _require(event.get("aggregate_type") == "risk_assessment", "FMEA_OUTBOX_BINDING_INVALID")


def _public_row(row: dict[str, object]) -> dict[str, object]:
    result = dict(row)
    field_evidence = row.get("field_evidence")
    field_support = row.get("field_support")
    _require(isinstance(field_evidence, list) and isinstance(field_support, list), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    result["field_evidence"] = [
        {"field_key": item[0], "evidence_ids": item[1]}
        for item in sorted(field_evidence, key=lambda item: item[0])
        if isinstance(item, list) and len(item) == 2
    ]
    result["field_support"] = [
        {"field_key": item[0], "support_status": item[1]}
        for item in sorted(field_support, key=lambda item: item[0])
        if isinstance(item, list) and len(item) == 2
    ]
    _require(len(result["field_evidence"]) == len(field_evidence), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    _require(len(result["field_support"]) == len(field_support), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    result["row_hash"] = _hash_json(row)
    return result


def _public_risk(record: dict[str, object]) -> dict[str, object]:
    fields = (
        "assessment_id",
        "derived",
        "dimensions",
        "domain_pack_id",
        "domain_pack_version",
        "evidence_pack_id",
        "invalidated_reason",
        "proposal_id",
        "record_version",
        "row_id",
        "rule_pack_id",
        "rule_pack_version",
        "source_record_version",
        "status",
        "workspace_id",
    )
    _require(all(field in record for field in fields), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    result = {field: record[field] for field in fields}
    result["assessment_hash"] = _hash_json(record)
    dimensions = result["dimensions"]
    _require(isinstance(dimensions, list), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    result["dimensions"] = sorted(dimensions, key=lambda item: item["name"])
    result["confirmation_basis"] = {"proposal_id": record["proposal_id"]}
    return result


_PUBLIC_EDGE_FIELDS = (
    "analysis_id",
    "barrier_ids",
    "claim_status",
    "delay_ms",
    "direction",
    "edge_id",
    "evidence_ids",
    "evidence_pack_id",
    "evidence_support",
    "fault_tolerance_time_ms",
    "interface_variable",
    "is_cyclic",
    "is_external",
    "is_terminal",
    "is_unprocessed",
    "operating_modes",
    "path_length",
    "publication_status",
    "record_version",
    "relation_type",
    "response_time_ms",
    "review_status",
    "risk_priority",
    "source_entity_id",
    "target_entity_id",
    "threshold",
    "unit",
)


def _public_edge(edge: dict[str, object]) -> dict[str, object]:
    _require(all(field in edge for field in _PUBLIC_EDGE_FIELDS), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    return {field: edge[field] for field in _PUBLIC_EDGE_FIELDS}


def _public_graph(graph: dict[str, object]) -> dict[str, object]:
    fields = (
        "analysis_id",
        "analysis_record_version",
        "domain_pack_id",
        "domain_pack_version",
        "graph_revision_id",
        "record_version",
        "rule_pack_id",
        "rule_pack_version",
        "status",
        "topology_hash",
        "topology_snapshot_id",
        "workspace_id",
    )
    _require(all(field in graph for field in fields), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    result = {field: graph[field] for field in fields}
    result["row_lineage"] = []
    result["nodes"] = [
        {field: node[field] for field in ("node_id", "node_type", "operating_modes")}
        for node in graph["nodes"]
    ]
    result["edges"] = [_public_edge(edge) for edge in graph["edges"]]
    result["paths"] = []
    for path in graph["paths"]:
        result["paths"].append({
            "analysis_id": path["analysis_id"],
            "edges": [_public_edge(edge) for edge in path["edges"]],
            "is_cyclic": path["is_cyclic"],
            "path_id": path["path_id"],
            "path_length": path["path_length"],
            "requires_human_review": path["requires_human_review"],
            "source_entity_id": path["source_entity_id"],
            "target_entity_id": path["target_entity_id"],
        })
    return result


def _public_locator(locator: object) -> dict[str, object]:
    if isinstance(locator, dict):
        return dict(locator)
    _require(isinstance(locator, str), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    match = re.fullmatch(r"page:(\d+)#span:(\d+)", locator)
    _require(match is not None, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    return {"page": int(match.group(1)), "span": int(match.group(2))}


def _public_evidence(packs: list[dict[str, object]]) -> list[dict[str, object]]:
    result = []
    for pack in sorted(packs, key=lambda item: item["pack_id"]):
        versions = pack.get("versions")
        _require(isinstance(versions, dict), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
        refs = []
        for ref in sorted(pack["refs"], key=lambda item: item["evidence_id"]):
            fields = ("content_hash", "document_id", "document_version", "evidence_hash", "evidence_id", "quote", "source_trust", "source_type")
            _require(all(field in ref for field in fields), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
            refs.append({
                **{field: ref[field] for field in fields},
                "locator": _public_locator(ref["locator"]),
            })
        result.append({
            "evidence_pack_version": versions.get("evidence_pack_version"),
            "pack_hash": _digest(pack["pack_hash"]),
            "pack_id": pack["pack_id"],
            "refs": refs,
        })
    return result


def _public_decisions(case: dict[str, object], candidates: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    audits = case["audits"]
    result = []
    for decision in case["review_decisions"]:
        row = decision.get("row") if isinstance(decision.get("row"), dict) else decision
        _require(isinstance(row, dict) and row.get("row_id") in candidates, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
        candidate = candidates[row["row_id"]]
        _require(row == candidate and decision.get("record_version") == candidate.get("record_version"), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
        matching = [
            audit for audit in audits
            if audit.get("command") == "review.decision"
            and audit.get("decision_id") == decision.get("decision_id")
            and audit.get("row_id") == row.get("row_id")
        ]
        _require(len(matching) == 1, "FMEA_REVIEW_RECEIPT_BINDING_INVALID")
        audit = matching[0]
        _require(audit.get("actor_type") == "human", "FMEA_AUTHORITY_BINDING_INVALID")
        _require(audit.get("applied_record_version") == candidate.get("record_version"), "FMEA_REVIEW_RECEIPT_BINDING_INVALID")
        _digest(audit.get("after_hash"))
        result.append({
            "analysis_id": candidate["analysis_id"],
            "decided_at": audit["occurred_at_server"],
            "decision": "accepted",
            "decision_id": decision["decision_id"],
            "reason": audit["reason"],
            "record_type": "row_review",
            "record_version": candidate["record_version"],
            "role_category": "human_reviewer",
            "row_hash": _hash_json(candidate),
            "row_id": candidate["row_id"],
            "workspace_id": audit["workspace_id"],
        })
    return sorted(result, key=lambda item: item["decision_id"])


def _template_bindings(case: dict[str, object], revisions: dict[str, dict[str, object]]) -> dict[tuple[str, str], dict[str, str]]:
    records = {}
    for item in case.get("registered_templates", []):
        compiled = item.get("compiled", {})
        metadata = compiled.get("metadata", {}) if isinstance(compiled, dict) else {}
        template_id = item.get("template_id") or metadata.get("template_id") or metadata.get("id")
        version = item.get("version") or metadata.get("version")
        template_hash = item.get("template_hash") or compiled.get("template_hash") if isinstance(compiled, dict) else None
        canonical = compiled.get("canonical_json") if isinstance(compiled, dict) else None
        _require(isinstance(template_id, str) and isinstance(version, str), "FMEA_TEMPLATE_BINDING_INVALID")
        _require(isinstance(template_hash, str) and isinstance(canonical, str), "FMEA_TEMPLATE_BINDING_INVALID")
        _require(_digest(template_hash) == sha256(canonical.encode("utf-8")).hexdigest(), "FMEA_TEMPLATE_CONTENT_MISMATCH")
        _require((template_id, version) not in records, "FMEA_DUPLICATE_RECORD")
        if isinstance(compiled, dict) and isinstance(compiled.get("template_hash"), str):
            _require(_digest(compiled["template_hash"]) == _digest(template_hash), "FMEA_TEMPLATE_CONTENT_MISMATCH")
        if "source_hash" in item:
            _digest(item["source_hash"])
        records[(template_id, version)] = {"template_hash": _digest(template_hash), "canonical_json": canonical}
    identities = set()
    for revision in revisions.values():
        for identity in revision["template_identities"]:
            _require(isinstance(identity, list) and len(identity) == 3, "FMEA_TEMPLATE_BINDING_INVALID")
            key = (identity[0], identity[1])
            identities.add(key)
            _require(records.get(key, {}).get("template_hash") == _digest(identity[2]), "FMEA_TEMPLATE_BINDING_INVALID")
    _require(identities <= set(records), "FMEA_TEMPLATE_BINDING_INVALID")
    return records


def _layout_value_path(field_key: str) -> list[str]:
    _require(_LAYOUT_FIELD_KEY.fullmatch(field_key) is not None and ".." not in field_key, "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    if field_key in _LAYOUT_ROW_FIELDS:
        return ["row", field_key]
    if "." in field_key:
        return ["extension_values", field_key]
    return ["unavailable", field_key]


def _independent_report_layout(canonical_json: str, identity: list[str]) -> dict[str, object]:
    source = _parse(canonical_json.encode("utf-8"))
    template = source.get("template")
    schema = source.get("output_schema")
    _require(isinstance(template, dict) and isinstance(schema, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    _require(
        template.get("id") == identity[0] and template.get("version") == identity[1],
        "FMEA_PUBLICATION_LAYOUT_MISMATCH",
    )
    properties = schema.get("properties", {})
    _require(isinstance(properties, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    properties = dict(properties)
    properties.setdefault("failure_mode", {"type": "string"})
    properties.setdefault("causes", {"type": "array", "items": {"type": "string"}})
    properties.setdefault("effects", {"type": "array", "items": {"type": "string"}})
    columns = []
    for field_key in sorted(properties):
        definition = properties[field_key]
        if isinstance(definition, bool):
            definition = {}
        _require(isinstance(definition, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
        value_type = definition.get("type", "json")
        items = definition.get("items", {})
        if value_type == "array" and isinstance(items, dict) and items.get("type") == "string":
            value_type = "string[]"
        if not isinstance(value_type, str) or value_type not in _LAYOUT_VALUE_TYPES:
            value_type = "json"
        label = definition.get("title", field_key)
        _require(isinstance(label, str), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
        columns.append(
            {
                "field_key": field_key,
                "label": label,
                "value_type": value_type,
                "value_path": _layout_value_path(field_key),
            }
        )
    return {
        "template_identity": {
            "template_id": identity[0],
            "version": identity[1],
            "template_hash": identity[2],
        },
        "columns": columns,
    }


def _select_independent_report_identity(
    revision: dict[str, object], templates: dict[tuple[str, str], dict[str, str]]
) -> list[str]:
    identities = revision.get("template_identities")
    _require(isinstance(identities, list) and identities, "FMEA_TEMPLATE_BINDING_INVALID")
    seen: set[tuple[str, str]] = set()
    direct_core: list[list[str]] = []
    for identity in identities:
        _require(isinstance(identity, list) and len(identity) == 3, "FMEA_TEMPLATE_BINDING_INVALID")
        _require(all(isinstance(value, str) for value in identity), "FMEA_TEMPLATE_BINDING_INVALID")
        key = (identity[0], identity[1])
        _require(key not in seen, "FMEA_DUPLICATE_RECORD")
        seen.add(key)
        template = templates.get(key)
        _require(template is not None, "FMEA_TEMPLATE_BINDING_INVALID")
        _require(template.get("template_hash") == _digest(identity[2]), "FMEA_TEMPLATE_BINDING_INVALID")
        source = _parse(template["canonical_json"].encode("utf-8"))
        schema = source.get("output_schema")
        _require(isinstance(schema, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
        properties = schema.get("properties")
        _require(isinstance(properties, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
        if {"failure_mode", "effects"}.issubset(properties):
            direct_core.append(identity)
    _require(len(seen) == len(identities), "FMEA_TEMPLATE_BINDING_INVALID")
    _require(len(direct_core) == 1, "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    return direct_core[0]


def _validate_publication_manifest_body(case: dict[str, object], snapshot: dict[str, object], revision: dict[str, object], templates: dict[tuple[str, str], dict[str, str]]) -> None:
    version_manifest = snapshot.get("version_manifest")
    _require(isinstance(version_manifest, dict), "FMEA_PUBLICATION_BODY_MARKER_MISSING")
    _require(version_manifest.get("body_schema_version") == BODY_SCHEMA_VERSION, "FMEA_PUBLICATION_BODY_MARKER_MISSING")
    expected_keys = {
        "analysis_hash",
        "body_schema_version",
        "domain_pack_identity",
        "propagation_rule_identity",
        "report_layout",
        "retrieval_provenance",
        "scoring_rule_identities",
        "template_identities",
    }
    _require(set(version_manifest) == expected_keys, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    for field in ("analysis_hash", "domain_pack_identity", "propagation_rule_identity", "scoring_rule_identities", "template_identities"):
        _require(version_manifest.get(field) == revision.get(field), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    layout = version_manifest.get("report_layout")
    _require(isinstance(layout, dict) and set(layout) == {"columns", "template_identity"}, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    report_identity = layout.get("template_identity")
    _require(isinstance(report_identity, dict) and set(report_identity) == {"template_hash", "template_id", "version"}, "FMEA_TEMPLATE_BINDING_INVALID")
    identity = [report_identity["template_id"], report_identity["version"], report_identity["template_hash"]]
    expected_identity = _select_independent_report_identity(revision, templates)
    _require(identity == expected_identity, "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    template = templates[(expected_identity[0], expected_identity[1])]
    _require(
        layout == _independent_report_layout(template["canonical_json"], expected_identity),
        "FMEA_PUBLICATION_LAYOUT_MISMATCH",
    )
    columns = layout.get("columns")
    _require(isinstance(columns, list) and bool(columns), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    for column in columns:
        _require(isinstance(column, dict) and set(column) == {"field_key", "label", "value_path", "value_type"}, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
        _require(isinstance(column["value_path"], list) and all(isinstance(item, str) for item in column["value_path"]), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    selected = case["evidence_selection"]
    native_provenance = revision.get("retrieval_provenance")
    _require(isinstance(selected, dict) and isinstance(native_provenance, dict), "FMEA_EVIDENCE_PROFILE_MISMATCH")
    profile_fields = ("requested_profile", "resolved_profile", "evidence_types", "source_counts", "warnings")
    _require(
        all(selected.get(field) == native_provenance.get(field) for field in profile_fields[:2]),
        "FMEA_EVIDENCE_PROFILE_MISMATCH",
    )
    provenance = version_manifest["retrieval_provenance"]
    _require(isinstance(provenance, dict), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    expected_provenance = {field: native_provenance.get(field) for field in profile_fields}
    _require(provenance == expected_provenance, "FMEA_EVIDENCE_PROFILE_MISMATCH")


def _validate_v2_publication_body(case: dict[str, object], revisions: dict[str, dict[str, object]]) -> None:
    candidates = _indexed(case["candidates"], "row_id")
    risks = sorted((record for record in case["risk_records"] if record.get("status") == "confirmed"), key=lambda item: item["assessment_id"])
    graphs = [graph for graph in case["propagation_graphs"] if graph.get("status") == "confirmed"]
    _require(len(graphs) == 1, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
    expected = {
        "rows": [_public_row(candidates[key]) for key in sorted(candidates)],
        "risk_records": [_public_risk(record) for record in risks],
        "propagation": _public_graph(graphs[0]),
        "evidence_summary": _public_evidence(case["evidence_packs"]),
        "decision_summary": _public_decisions(case, candidates),
        "unresolved_items": [],
    }
    templates = _template_bindings(case, revisions)
    for snapshot in case["snapshots"]:
        revision = revisions.get(snapshot.get("revision_id"))
        _require(revision is not None, "FMEA_SNAPSHOT_CONTENT_MISMATCH")
        _validate_publication_manifest_body(case, snapshot, revision, templates)
        for field, value in expected.items():
            _require(snapshot.get(field) == value, "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")
        _require(snapshot.get("row_count") == len(expected["rows"]), "FMEA_PUBLICATION_BODY_NATIVE_MISMATCH")


def _visible_plain(value: object) -> object:
    if isinstance(value, dict):
        return {str(key): _visible_plain(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_visible_plain(item) for item in value]
    return value


def _visible_text(value: object) -> str:
    if value is None:
        return "（无）"
    if isinstance(value, str):
        if value == "":
            return "（空字符串）"
        if not value.strip():
            return "（空白字符串）"
        return value
    if isinstance(value, dict):
        if not value:
            return "（空对象）"
        return json.dumps(_visible_plain(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if isinstance(value, list):
        if not value:
            return "（空列表）"
        return "；".join(_visible_text(item) for item in value)
    return str(value)


def _visible_row_values(row: dict[str, object], layout: dict[str, object]) -> list[str]:
    extensions = {
        item["field_key"]: item.get("value")
        for item in row.get("extension_values", [])
        if isinstance(item, dict) and isinstance(item.get("field_key"), str)
    }
    values = []
    for column in layout["columns"]:
        path = column["value_path"]
        _require(isinstance(path, list) and len(path) == 2, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        kind, key = path
        value = row.get(key) if kind == "row" else extensions.get(key) if kind == "extension_values" else None
        values.append(_visible_text(value))
    return values


def _visible_risk_summary(snapshot: dict[str, object], row: dict[str, object]) -> str:
    risks = [
        risk
        for risk in snapshot["risk_records"]
        if risk.get("row_id") == row.get("row_id")
        and risk.get("source_record_version") == row.get("record_version")
    ]
    summaries = []
    for risk in risks:
        parts = [str(risk["assessment_id"]), str(risk["status"])]
        derived = risk.get("derived")
        if isinstance(derived, dict):
            for key in ("rpn", "priority"):
                if key in derived:
                    parts.append(f"{key}={_visible_text(derived[key])}")
        summaries.append("；".join(parts))
    return " | ".join(summaries)


def _visible_evidence_ids(row: dict[str, object]) -> str:
    identifiers = set()
    for binding in row.get("field_evidence", []):
        if isinstance(binding, dict):
            identifiers.update(str(item) for item in binding.get("evidence_ids", []))
    return "、".join(sorted(identifiers))


def _visible_main_values(snapshot: dict[str, object], layout: dict[str, object]) -> list[list[str]]:
    result = []
    for row in snapshot["rows"]:
        values = _visible_row_values(row, layout)
        values.extend((_visible_risk_summary(snapshot, row), _visible_text(row.get("review_status")), _visible_evidence_ids(row)))
        result.append(values)
    return result


def _visible_detail_append(result: list[tuple[str, str, str, str, str]], row_id: str, version: str, detail_type: str, prefix: str, value: object) -> None:
    if isinstance(value, dict):
        if not value:
            result.append((row_id, version, detail_type, prefix, "（空对象）"))
            return
        for key, item in value.items():
            field = f"{prefix}.{key}" if prefix else str(key)
            _visible_detail_append(result, row_id, version, detail_type, field, item)
        return
    if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
        for index, item in enumerate(value, start=1):
            _visible_detail_append(result, row_id, version, detail_type, f"{prefix}[{index}]", item)
        return
    result.append((row_id, version, detail_type, prefix, _visible_text(value)))


def _visible_detail_rows(snapshot: dict[str, object]) -> list[tuple[str, str, str, str, str]]:  # noqa: C901 - independent detail semantics
    result: list[tuple[str, str, str, str, str]] = []
    for row in snapshot["rows"]:
        row_id = _visible_text(row.get("row_id"))
        version = _visible_text(row.get("record_version"))
        for key, value in row.items():
            if key in {"extension_values", "field_evidence", "field_support", "field_claims"}:
                continue
            result.append((row_id, version, "正文", str(key), _visible_text(value)))
        for extension in row.get("extension_values", []):
            if isinstance(extension, dict):
                result.append((row_id, version, "扩展字段", _visible_text(extension.get("field_key")), _visible_text(extension.get("value"))))
        for field, detail_type in (("field_evidence", "证据绑定"), ("field_support", "字段支持"), ("field_claims", "字段声明")):
            if row.get(field):
                result.append((row_id, version, detail_type, field, _visible_text(row[field])))
        for risk in snapshot["risk_records"]:
            if risk.get("row_id") == row.get("row_id") and risk.get("source_record_version") == row.get("record_version"):
                _visible_detail_append(result, row_id, version, "评分", _visible_text(risk.get("assessment_id")), risk)
        for decision in snapshot["decision_summary"]:
            if isinstance(decision, dict) and decision.get("row_id") == row.get("row_id"):
                _visible_detail_append(result, row_id, version, "复核", _visible_text(decision.get("decision_id")), decision)

    seen_packs: set[str] = set()
    seen_evidence: set[str] = set()
    global_version = _visible_text("")
    for pack in snapshot["evidence_summary"]:
        if not isinstance(pack, dict):
            continue
        pack_id = _visible_text(pack.get("pack_id"))
        if pack_id not in seen_packs:
            seen_packs.add(pack_id)
            _visible_detail_append(result, "", global_version, "共享证据包", f"pack.{pack_id}", {key: value for key, value in pack.items() if key != "refs"})
        for reference in pack.get("refs", []):
            if not isinstance(reference, dict):
                continue
            identity = reference.get("evidence_id")
            identity = identity if isinstance(identity, str) else _visible_text(reference)
            if identity in seen_evidence:
                continue
            seen_evidence.add(identity)
            _visible_detail_append(result, "", global_version, "证据", identity, reference)
    if not snapshot["rows"]:
        for decision in snapshot["decision_summary"]:
            if isinstance(decision, dict):
                _visible_detail_append(result, "", global_version, "复核", "decision", decision)
    if snapshot.get("propagation") is not None:
        _visible_detail_append(result, "", global_version, "传播", "propagation", snapshot["propagation"])
    return result


def _collapse_visible_detail_rows(rows: list[tuple[str, str, str, str, str]]) -> list[tuple[str, str, str, str, str]]:
    segment_pattern = re.compile(r"(.+) \[part ([1-9][0-9]*)/([1-9][0-9]*)\]\Z")
    order: list[tuple[str, object]] = []
    segments: dict[tuple[str, str, str, str], tuple[int, dict[int, str]]] = {}
    for row_id, version, detail_type, field, content in rows:
        match = segment_pattern.fullmatch(field)
        if match is None:
            order.append(("plain", (row_id, version, detail_type, field, content)))
            continue
        base = match.group(1)
        part = int(match.group(2))
        total = int(match.group(3))
        key = (row_id, version, detail_type, base)
        if key not in segments:
            segments[key] = (total, {})
            order.append(("segment", key))
        expected_total, values = segments[key]
        _require(expected_total == total and part not in values, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        values[part] = content
    result = []
    for kind, value in order:
        if kind == "plain":
            result.append(value)
            continue
        row_id, version, detail_type, field = value
        total, values = segments[value]
        _require(set(values) == set(range(1, total + 1)), "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        result.append((row_id, version, detail_type, field, "".join(values[index] for index in range(1, total + 1))))
    return result


def _visible_xlsx_body(payload: bytes, snapshot: dict[str, object], layout: dict[str, object]) -> None:
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    try:
        main_rows = [tuple("" if value is None else str(value) for value in row) for row in workbook["正文"].iter_rows(values_only=True)]
        expected_headers = ["逻辑行", "分段", "续字段", *[column["label"] for column in layout["columns"]], "评分摘要", "复核状态", "证据编号"]
        _require(main_rows and list(main_rows[0]) == expected_headers, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        actual_rows = [row for row in main_rows[1:] if any(value != "" for value in row)]
        expected_values = _visible_main_values(snapshot, layout)
        cursor = 0
        for logical_index, values in enumerate(expected_values, start=1):
            group = []
            while cursor < len(actual_rows) and actual_rows[cursor][0] == f"第{logical_index}行":
                group.append(actual_rows[cursor])
                cursor += 1
            _require(group, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
            parts = []
            for row in group:
                match = re.fullmatch(r"([1-9][0-9]*)/([1-9][0-9]*)", row[1])
                _require(match is not None, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
                parts.append((int(match.group(1)), int(match.group(2))))
            total = parts[0][1]
            _require(total == len(group) and sorted(part for part, _ in parts) == list(range(1, total + 1)), "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
            continued = [
                field
                for field_index, field in enumerate([column["field_key"] for column in layout["columns"]] + ["risk_summary", "review_status", "evidence_ids"], start=3)
                if any(row[field_index] != "" for row in group[1:])
            ]
            expected_continued = "、".join(continued)
            _require(all(row[2] == expected_continued for row in group), "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
            for field_index, expected in enumerate(values, start=3):
                _require("".join(row[field_index] for row in group) == expected, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        _require(cursor == len(actual_rows), "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        detail_sheet = workbook["正文详情"]
        _require(detail_sheet.max_column == 5, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        detail_table = [
            tuple("" if value is None else str(value) for value in row)
            for row in detail_sheet.iter_rows(values_only=True)
            if any(value is not None and value != "" for value in row)
        ]
        _require(all(len(row) == 5 for row in detail_table), "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        _require(detail_table and list(detail_table[0]) == ["行ID", "记录版本", "详情类型", "字段", "内容"], "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
        _require(
            _collapse_visible_detail_rows(detail_table[1:]) == _visible_detail_rows(snapshot),
            "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH",
        )
    finally:
        workbook.close()


def _visible_docx_body(payload: bytes, snapshot: dict[str, object], layout: dict[str, object]) -> None:  # noqa: C901 - independent visible-body semantics
    members = _office_members(payload)
    root = safe_xml_fromstring(members["word/document.xml"], forbid_dtd=True, forbid_entities=True, forbid_external=True)
    body = root.find(f"{_WORD}body")
    _require(body is not None, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
    paragraphs = []
    tables = []
    for child in body:
        if child.tag == f"{_WORD}p":
            text = _word_paragraph(child)
            if text.startswith("Canonical table: "):
                break
            if text:
                paragraphs.append(text)
        elif child.tag == f"{_WORD}tbl":
            tables.append([
                [_word_cell(cell) for cell in row.findall(f"{_WORD}tc")]
                for row in child.findall(f"{_WORD}tr")
            ])
    expected_values = _visible_main_values(snapshot, layout)
    expected_table = [[column["label"] for column in layout["columns"][:3]]]
    expected_table.extend([values[:3] for values in expected_values])
    _require(tables == [expected_table], "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")
    expected_paragraphs = ["FMEA Export", "FMEA 正文", "逐行详情"]
    declared_keys = {column["field_key"] for column in layout["columns"]}
    for row, values in zip(snapshot["rows"], expected_values, strict=True):
        row_id = _visible_text(row.get("row_id"))
        version = _visible_text(row.get("record_version"))
        expected_paragraphs.append(f"行ID：{row_id}（记录版本：{version}）")
        for column, value in zip(layout["columns"], values[: len(layout["columns"])], strict=True):
            expected_paragraphs.append(f"{column['label']} [{column['field_key']}]：{value}")
        for key, value in row.items():
            if key == "extension_values" or key in declared_keys:
                continue
            expected_paragraphs.append(f"{key}：{_visible_text(value)}")
        for extension in row.get("extension_values", []):
            if isinstance(extension, dict):
                expected_paragraphs.append(
                    f"扩展字段 {extension.get('field_key')}（{extension.get('value_type')}）：{_visible_text(extension.get('value'))}"
                )
        for risk in snapshot["risk_records"]:
            if risk.get("row_id") == row.get("row_id") and risk.get("source_record_version") == row.get("record_version"):
                _visible_docx_mapping_paragraphs(expected_paragraphs, f"评分 {_visible_text(risk.get('assessment_id'))}", risk)
        for decision in snapshot["decision_summary"]:
            if isinstance(decision, dict) and decision.get("row_id") == row.get("row_id"):
                _visible_docx_mapping_paragraphs(expected_paragraphs, f"复核 {_visible_text(decision.get('decision_id'))}", decision)
    expected_paragraphs.append("共享证据")
    seen_packs: set[str] = set()
    seen_evidence: set[str] = set()
    for pack in snapshot["evidence_summary"]:
        if not isinstance(pack, dict):
            continue
        pack_id = _visible_text(pack.get("pack_id"))
        if pack_id not in seen_packs:
            seen_packs.add(pack_id)
            _visible_docx_mapping_paragraphs(expected_paragraphs, f"证据包 {pack_id}", {key: value for key, value in pack.items() if key != "refs"})
        for reference in pack.get("refs", []):
            if not isinstance(reference, dict):
                continue
            identity = reference.get("evidence_id")
            identity = identity if isinstance(identity, str) else _visible_text(reference)
            if identity in seen_evidence:
                continue
            seen_evidence.add(identity)
            _visible_docx_mapping_paragraphs(expected_paragraphs, f"证据 {identity}", reference)
    if not snapshot["rows"]:
        for decision in snapshot["decision_summary"]:
            if isinstance(decision, dict):
                _visible_docx_mapping_paragraphs(expected_paragraphs, f"复核 {_visible_text(decision.get('decision_id'))}", decision)
    if snapshot.get("propagation") is not None:
        expected_paragraphs.append("传播")
        _visible_docx_mapping_paragraphs(expected_paragraphs, "传播", snapshot["propagation"])
    expected_paragraphs.append("机器附录")
    _require(paragraphs == expected_paragraphs, "FMEA_PUBLICATION_VISIBLE_BODY_MISMATCH")


def _visible_docx_mapping_paragraphs(result: list[str], prefix: str, value: object) -> None:
    if isinstance(value, dict):
        if not value:
            result.append(f"{prefix}：（空对象）")
            return
        for key, item in value.items():
            field = f"{prefix}.{key}" if prefix else str(key)
            _visible_docx_mapping_paragraphs(result, field, item)
        return
    if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
        for index, item in enumerate(value, start=1):
            _visible_docx_mapping_paragraphs(result, f"{prefix}[{index}]", item)
        return
    result.append(f"{prefix}：{_visible_text(value)}")


def _validate_visible_office_body(payload: bytes, format_name: str, snapshot: dict[str, object]) -> None:
    layout = snapshot.get("version_manifest", {}).get("report_layout")
    _require(isinstance(layout, dict), "FMEA_PUBLICATION_LAYOUT_MISMATCH")
    if format_name == "xlsx":
        _visible_xlsx_body(payload, snapshot, layout)
    elif format_name == "docx":
        _visible_docx_body(payload, snapshot, layout)


def _validate_versioned_content(case: dict, revisions: dict, migration_child: dict, contract_version: str = SCHEMA_VERSION) -> None:
    if contract_version == LEGACY_SCHEMA_VERSION:
        _validate_legacy_versioned_content(case, revisions, migration_child)
        return
    rows = sorted(case["candidates"], key=lambda row: row["row_id"])
    risks = sorted((record for record in case["risk_records"] if record["status"] == "confirmed"), key=lambda record: record["assessment_id"])
    row_versions = [[row["row_id"], row["record_version"], _hash_json(row)] for row in rows]
    risk_versions = [[risk["assessment_id"], risk["record_version"], _hash_json(risk)] for risk in risks]
    graph = next(item for item in case["propagation_graphs"] if item["status"] == "confirmed")
    for revision in revisions.values():
        _require(sorted(revision["row_versions"]) == row_versions, "FMEA_REVISION_CONTENT_MISMATCH")
        if revision["revision_id"] == migration_child["revision_id"]:
            _require(revision["risk_versions"] == [] and revision.get("propagation_graph_revision_id") is None and revision.get("propagation_graph_hash") is None, "FMEA_MIGRATION_BINDING_INVALID")
        else:
            _require(sorted(revision["risk_versions"]) == risk_versions, "FMEA_REVISION_CONTENT_MISMATCH")
            _require(revision.get("propagation_graph_revision_id") == graph["graph_revision_id"] and _digest(revision.get("propagation_graph_hash")) == _hash_json(graph), "FMEA_PROPAGATION_CONTENT_MISMATCH")
    for risk in risks:
        event = _one_event(case, "risk.confirmed", risk["assessment_id"])
        _require(event.get("payload", {}).get("assessment") == risk, "FMEA_RISK_CONTENT_MISMATCH")
        _require(event.get("aggregate_type") == "risk_assessment", "FMEA_OUTBOX_BINDING_INVALID")
    _validate_v2_publication_body(case, revisions)


def _topology_body(topology: dict) -> dict:
    body = {key: topology[key] for key in ("workspace_id", "analysis_id", "record_version", "created_at")}
    body["id"] = topology["topology_snapshot_id"]
    body["nodes"] = [{"id": node["node_id"], "type": node["node_type"], "operating_modes": node["operating_modes"]} for node in topology["nodes"]]
    body["interfaces"] = [{"id": item["interface_id"], **{key: item[key] for key in ("source_node_id", "target_node_id", "interface_variable", "unit", "direction", "operating_modes")}} for item in topology["interfaces"]]
    return body


def _rule_hash(rule: dict) -> str:
    body = {key: value for key, value in rule.items() if key != "rule_pack_id"}
    body["id"] = rule["rule_pack_id"]
    body["timing_constraints"] = dict.fromkeys(("delay_ms", "response_time_ms", "fault_tolerance_time_ms"), "non_negative")
    return _hash_json(body)


def _validate_graph_edge_structure(edge: dict, graph: dict, topology: dict, rule: dict, nodes: dict) -> None:
    code = "FMEA_PROPAGATION_TOPOLOGY_INVALID"
    source, target = edge.get("source_entity_id"), edge.get("target_entity_id")
    _require(source in nodes and target in nodes, code)
    _require(edge.get("analysis_id") == graph["analysis_id"], code)
    for field, allowed in (("relation_type", "relation_types"), ("interface_variable", "interface_variables"), ("unit", "units"), ("direction", "directions")):
        _require(edge.get(field) in rule.get(allowed, []), code)
    modes = set(edge.get("operating_modes", []))
    matches = [item for item in topology["interfaces"] if item["source_node_id"] == source and item["target_node_id"] == target and all(item[field] == edge[field] for field in ("interface_variable", "unit", "direction"))]
    _require(bool(matches) and all(modes.intersection(item["operating_modes"]) for item in matches), code)
    _require(all(not nodes[node]["operating_modes"] or modes.intersection(nodes[node]["operating_modes"]) for node in (source, target)), code)
    _require(all(type(edge.get(field)) is int and edge[field] >= 0 for field in ("delay_ms", "response_time_ms", "fault_tolerance_time_ms")), code)
    _require(edge.get("review_status") == ("accepted" if graph["status"] == "confirmed" else "suggested"), "FMEA_PROPAGATION_REVIEW_INVALID")


def _validate_graph_paths(graph: dict, rule: dict, edges: dict) -> None:
    code = "FMEA_PROPAGATION_PATH_INVALID"
    paths = _indexed(_required_list(graph, "paths"), "path_id")
    for path in paths.values():
        path_edges = path.get("edges")
        _require(isinstance(path_edges, list) and bool(path_edges) and path.get("path_length") == len(path_edges), code)
        _require(all(edges.get(edge.get("edge_id")) == edge for edge in path_edges), code)
        _require(path.get("source_entity_id") == path_edges[0]["source_entity_id"] and path.get("target_entity_id") == path_edges[-1]["target_entity_id"], code)
        _require(all(left["target_entity_id"] == right["source_entity_id"] for left, right in pairwise(path_edges)), code)
        _require(path.get("analysis_id") == graph["analysis_id"], code)
        visited = [edge["source_entity_id"] for edge in path_edges] + [path_edges[-1]["target_entity_id"]]
        cyclic = len(visited) != len(set(visited)) or any(edge.get("is_cyclic") for edge in path_edges)
        _require(path.get("is_cyclic") is cyclic, code)
        requires_review = cyclic or len(path_edges) > rule["max_automatic_depth"] or any(edge.get("risk_priority") in {"high", "critical"} or edge.get("is_external") or edge.get("is_unprocessed") or edge.get("barrier_ids") for edge in path_edges)
        _require(not requires_review or path.get("requires_human_review") is True, code)


def _validate_graph_structure(case: dict, revisions: dict) -> None:
    topologies = _indexed(_required_list(case, "topology_snapshots"), "topology_snapshot_id")
    rules = {(rule["rule_pack_id"], rule["version"]): rule for rule in _required_list(case, "rule_packs")}
    _require(len(rules) == len(case["rule_packs"]), "FMEA_DUPLICATE_RECORD")
    for graph in case["propagation_graphs"]:
        topology = topologies.get(graph["topology_snapshot_id"])
        rule = rules.get((graph["rule_pack_id"], graph["rule_pack_version"]))
        _require(topology is not None and rule is not None, "FMEA_PROPAGATION_TOPOLOGY_INVALID")
        _require(_hash_json(_topology_body(topology)) == _digest(topology["topology_hash"]) == _digest(graph["topology_hash"]), "FMEA_TOPOLOGY_HASH_MISMATCH")
        _require(topology["workspace_id"] == graph["workspace_id"] and topology["analysis_id"] == graph["analysis_id"], "FMEA_PROPAGATION_TOPOLOGY_INVALID")
        _require(rule.get("max_automatic_depth") == 2 and rule.get("prohibit_silent_fallback") is True, "FMEA_PROPAGATION_RULE_INVALID")
        rule_identity = [rule["rule_pack_id"], rule["version"], _rule_hash(rule)]
        _require(all(revision.get("propagation_rule_identity") == rule_identity for revision in revisions.values()), "FMEA_PROPAGATION_RULE_INVALID")
        nodes = _indexed(topology["nodes"], "node_id")
        graph_nodes = _indexed(graph["nodes"], "node_id")
        _require(all(nodes.get(identity) == node for identity, node in graph_nodes.items()), "FMEA_PROPAGATION_TOPOLOGY_INVALID")
        edges = _indexed(graph["edges"], "edge_id")
        for edge in edges.values():
            _validate_graph_edge_structure(edge, graph, topology, rule, nodes)
        _validate_graph_paths(graph, rule, edges)
        event = _one_event(case, f"propagation.{graph['status']}", graph["graph_revision_id"])
        _require(event.get("payload", {}).get("graph") == graph, "FMEA_PROPAGATION_CONTENT_MISMATCH")


def _manifest_version_hash(revision: dict[str, object], body_schema_version: str | None = None) -> str:
    body = {
        "revision_hash": revision.get("revision_hash"),
        "analysis_hash": revision.get("analysis_hash"),
        "domain_pack_identity": revision.get("domain_pack_identity"),
        "template_identities": revision.get("template_identities"),
        "scoring_rule_identities": revision.get("scoring_rule_identities"),
        "propagation_rule_identity": revision.get("propagation_rule_identity"),
    }
    if body_schema_version is not None:
        body["body_schema_version"] = body_schema_version
    return _hash_json(body, ascii_only=False)


def _validate_manifests(
    case: dict[str, object],
    revisions: dict[str, dict[str, object]],
    publications: dict[str, dict[str, object]],
    contract_version: str = SCHEMA_VERSION,
) -> None:
    manifests = _indexed(case["manifests"], "manifest_id")
    snapshots = _indexed(case["snapshots"], "snapshot_id")
    published_revision_ids = {publication.get("revision_id") for publication in publications.values()}
    _require({manifest.get("revision_id") for manifest in manifests.values()} == published_revision_ids, "FMEA_MANIFEST_BINDING_INVALID")
    for manifest in manifests.values():
        revision = revisions.get(manifest.get("revision_id"))
        _require(revision is not None and manifest.get("export_eligible") is True, "FMEA_MANIFEST_BINDING_INVALID")
        _require(manifest.get("revision_hash") == revision.get("revision_hash"), "FMEA_MANIFEST_BINDING_INVALID")
        _digest(manifest.get("manifest_hash"))
        body = {key: manifest.get(key) for key in (
            "manifest_id", "revision_id", "revision_hash", "approval_id", "snapshot_id", "snapshot_hash",
            "version_manifest_hash", "previous_audit_chain_head", "export_eligible",
        )}
        _require(_hash_json(body, ascii_only=False) == _digest(manifest.get("manifest_hash")), "FMEA_MANIFEST_HASH_MISMATCH")
        if "version_manifest_hash" in manifest:
            snapshot = snapshots.get(manifest.get("snapshot_id"), {})
            marker = None
            if contract_version == SCHEMA_VERSION:
                marker = snapshot.get("version_manifest", {}).get("body_schema_version")
                _require(marker == BODY_SCHEMA_VERSION, "FMEA_PUBLICATION_BODY_MARKER_MISSING")
            _require(_digest(manifest["version_manifest_hash"]) == _digest(_manifest_version_hash(revision, marker)), "FMEA_MANIFEST_BINDING_INVALID")


def _validate_submissions(case: dict[str, object], revisions: dict[str, dict[str, object]], approvals: dict[str, dict[str, object]]) -> None:
    submissions = _indexed(case["submissions"], "submission_id")
    _require(len(submissions) == len(approvals), "FMEA_APPROVAL_BINDING_INVALID")
    for approval in approvals.values():
        submission = submissions.get(approval.get("submission_id"))
        revision = revisions.get(approval.get("revision_id"))
        _require(submission is not None and revision is not None, "FMEA_APPROVAL_BINDING_INVALID")
        _require(submission.get("revision_id") == revision.get("revision_id"), "FMEA_APPROVAL_BINDING_INVALID")
        _require(submission.get("revision_hash") == revision.get("revision_hash"), "FMEA_APPROVAL_BINDING_INVALID")
        _require(submission.get("status") == "pending", "FMEA_APPROVAL_BINDING_INVALID")
        _require(submission.get("submitter_actor_id"), "FMEA_AUTHORITY_BINDING_INVALID")


def _validate_template_input(payload: bytes) -> None:
    """Apply the importer's intentionally strict plain-XLSX policy."""
    members = _office_members(payload)
    for name, content in members.items():
        if name.casefold().endswith(".xml"):
            root = safe_xml_fromstring(content, forbid_dtd=True, forbid_entities=True, forbid_external=True)
            _require(
                all(node.tag.rsplit("}", 1)[-1].casefold() != "definedname" for node in root.iter()),
                "FMEA_TEMPLATE_INPUT_INVALID",
            )
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        try:
            _require(len(workbook.sheetnames) == 1, "FMEA_TEMPLATE_INPUT_INVALID")
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
    except (BadZipFile, KeyError, OSError, ValueError, TypeError):
        raise VerificationError("FMEA_TEMPLATE_INPUT_INVALID") from None
    _require(bool(rows) and len(rows[0]) >= 2, "FMEA_TEMPLATE_INPUT_INVALID")
    headers = {str(value).strip() for value in rows[0] if value not in {None, ""}}
    _require({"failure_mode", "legacy_criticality"} <= headers, "FMEA_TEMPLATE_INPUT_INVALID")


def _validate_migration(case: dict[str, object], root: dict[str, object], migration_child: dict[str, object], exports: list[dict[str, object]], payloads: dict[str, bytes] | None) -> None:
    reports = _required_list(case, "migration_reports")
    _require(len(reports) == 1, "FMEA_MIGRATION_BINDING_INVALID")
    report = reports[0]
    _digest(report.get("report_hash"))
    _require(report.get("migration_id") == case["migration_results"][0].get("migration_id"), "FMEA_MIGRATION_BINDING_INVALID")
    _require(report.get("status") in {"confirmed", "dry_run"}, "FMEA_MIGRATION_BINDING_INVALID")
    _require(report.get("source_revision_id") == root.get("revision_id"), "FMEA_MIGRATION_BINDING_INVALID")
    _require(report.get("source_revision_hash") == root.get("revision_hash"), "FMEA_MIGRATION_BINDING_INVALID")
    _require(
        report.get("target_domain_pack_identity") == migration_child.get("domain_pack_identity")
        and report.get("source_domain_pack_identity") == root.get("domain_pack_identity"),
        "FMEA_MIGRATION_BINDING_INVALID",
    )
    completed = _one_event(case, "migration.completed", migration_child["revision_id"])
    body = completed.get("payload", {})
    for field in ("migration_id", "report_hash", "source_revision_id", "source_revision_hash", "source_domain_pack_identity", "target_domain_pack_identity", "target_revision_hash", "plan", "mapped_fields", "dropped_fields", "unresolved_fields", "warnings"):
        _require(body.get(field) == report.get(field), "FMEA_MIGRATION_BINDING_INVALID")
    _require(body.get("child_revision_id") == migration_child["revision_id"] and body.get("child_revision_hash") == migration_child["revision_hash"], "FMEA_MIGRATION_BINDING_INVALID")
    receipt = case.get("invalidation_receipt", {})
    _require(receipt.get("outbox_event_id") == completed["event_id"] and receipt.get("child_revision_id") == migration_child["revision_id"] and receipt.get("source_revision_id") == root["revision_id"], "FMEA_MIGRATION_BINDING_INVALID")
    _require(receipt.get("risk_versions_before") == root["risk_versions"] and receipt.get("propagation_graph_revision_id_before") == root["propagation_graph_revision_id"], "FMEA_MIGRATION_BINDING_INVALID")
    _require(receipt.get("risk_invalidated") is True and receipt.get("propagation_invalidated") is True, "FMEA_MIGRATION_BINDING_INVALID")
    plan = report.get("plan")
    _require(isinstance(plan, dict) and plan.get("source") and plan.get("target") and plan.get("steps"), "FMEA_MIGRATION_BINDING_INVALID")
    results = case["migration_results"]
    _composite_index(results, ("migration_id", "replayed"))
    _require(all("child_revision" not in result for result in results), "FMEA_MIGRATION_RESULT_SCHEMA_INVALID")
    for result in results:
        _require(result.get("migration_id") == report.get("migration_id"), "FMEA_MIGRATION_BINDING_INVALID")
        _require(result.get("child_revision_id") == migration_child.get("revision_id"), "FMEA_MIGRATION_BINDING_INVALID")
        _require(_digest(result.get("report_hash")) == _digest(report.get("report_hash")), "FMEA_MIGRATION_BINDING_INVALID")
        _require(type(result.get("replayed")) is bool, "FMEA_REPLAY_SCHEMA_INVALID")
    _require(any(result.get("replayed") is False for result in results), "FMEA_MIGRATION_BINDING_INVALID")
    drafts = _required_list(case, "template_drafts")
    _require(any(draft.get("status") == "draft" for draft in drafts), "FMEA_TEMPLATE_IMPORT_INVALID")
    sources = _required_list(case, "template_import_sources")
    _require(len(sources) == 1, "FMEA_TEMPLATE_IMPORT_INVALID")
    source = sources[0]
    source_path = _text(source.get("path"), "FMEA_TEMPLATE_IMPORT_INVALID")
    _require(source_path.startswith("inputs/") and source_path.endswith(".xlsx"), "FMEA_TEMPLATE_IMPORT_INVALID")
    source_hash = _digest(source.get("sha256"))
    _require(type(source.get("byte_length")) is int and source["byte_length"] > 0, "FMEA_TEMPLATE_IMPORT_INVALID")
    for draft in drafts:
        _digest(draft.get("source_sha256"))
        _require(draft.get("structure") or draft.get("source_structure"), "FMEA_TEMPLATE_IMPORT_INVALID")
        _require(_digest(draft["source_sha256"]) == source_hash, "FMEA_TEMPLATE_IMPORT_INVALID")
        if "source_filename" in draft:
            _require(draft["source_filename"] == source_path.rsplit("/", 1)[-1], "FMEA_TEMPLATE_IMPORT_INVALID")
    if payloads is not None:
        _require(source_path in payloads, "FMEA_TEMPLATE_IMPORT_INVALID")
        _require(len(payloads[source_path]) == source["byte_length"], "FMEA_TEMPLATE_IMPORT_INVALID")
        _require(sha256(payloads[source_path]).hexdigest() == source_hash, "FMEA_TEMPLATE_IMPORT_INVALID")
        _validate_template_input(payloads[source_path])


def _validate_export_wrappers(
    case: dict[str, object],
    publications: dict[str, dict[str, object]],
    payloads: dict[str, bytes] | None,
) -> None:
    exports = _indexed(case["exports"], "path", "FMEA_DUPLICATE_EXPORT")
    snapshots = _indexed(case["snapshots"], "snapshot_id")
    _require(len(exports) == 6, "FMEA_EXPORT_SET_INCOMPLETE")
    for path, export in exports.items():
        _require(set(export) >= {"path", "format", "run", "manifest"}, "FMEA_EXPORT_SCHEMA_INVALID")
        _require(export.get("path") == path, "FMEA_EXPORT_SCHEMA_INVALID")
        _require(export.get("format") in {"json", "xlsx", "docx"}, "FMEA_EXPORT_FORMAT_INVALID")
        run, manifest = export.get("run"), export.get("manifest")
        _require(isinstance(run, dict) and isinstance(manifest, dict), "FMEA_EXPORT_SCHEMA_INVALID")
        _require(run.get("status") == "succeeded" and run.get("draft_preview") is False, "FMEA_EXPORT_NOT_PUBLISHED")
        _require(run.get("format") == export.get("format"), "FMEA_EXPORT_FORMAT_INVALID")
        _require(manifest.get("export_run_id") == run.get("export_run_id"), "FMEA_EXPORT_BINDING_INVALID")
        _require(manifest.get("format") == export.get("format") and manifest.get("draft_preview") is False, "FMEA_EXPORT_BINDING_INVALID")
        snapshot_id = run.get("snapshot_id") or manifest.get("snapshot_id")
        _require(snapshot_id in snapshots, "FMEA_EXPORT_SNAPSHOT_UNBOUND")
        _require(manifest.get("snapshot_id") == snapshot_id, "FMEA_EXPORT_BINDING_INVALID")
        _require(manifest.get("revision_id") == snapshots[snapshot_id].get("revision_id"), "FMEA_EXPORT_BINDING_INVALID")
        publication_id = snapshots[snapshot_id].get("publication_id")
        _require(manifest.get("publication_id") == publication_id, "FMEA_EXPORT_BINDING_INVALID")
        _require(publication_id in publications, "FMEA_EXPORT_BINDING_INVALID")
        for field in ("snapshot_id", "snapshot_hash", "revision_id", "publication_id"):
            expected = snapshots[snapshot_id].get(field)
            _require(run.get(field) == expected and manifest.get(field) == expected, "FMEA_EXPORT_BINDING_INVALID")
        _require(run.get("workspace_id") == snapshots[snapshot_id].get("workspace_id"), "FMEA_EXPORT_BINDING_INVALID")
        _require(run.get("artifact_id") == manifest.get("artifact_id"), "FMEA_EXPORT_BINDING_INVALID")
        _require(run.get("filename") == manifest.get("filename") == path.rsplit("/", 1)[-1], "FMEA_EXPORT_BINDING_INVALID")
        _digest(manifest.get("sha256"))
        _require(type(manifest.get("byte_length")) is int and manifest["byte_length"] > 0, "FMEA_EXPORT_SCHEMA_INVALID")
        if payloads is not None:
            _require(path in payloads, "FMEA_EXPORT_FORMAT_INVALID")
            _require(len(payloads[path]) == manifest["byte_length"], "FMEA_EXPORT_HASH_MISMATCH")
            _require(sha256(payloads[path]).hexdigest() == _digest(manifest["sha256"]), "FMEA_EXPORT_HASH_MISMATCH")
    by_snapshot = {snapshot_id: set() for snapshot_id in snapshots}
    for export in exports.values():
        run = export["run"]
        by_snapshot[run.get("snapshot_id")].add(export["format"])
    _require(all(formats == {"json", "xlsx", "docx"} for formats in by_snapshot.values()), "FMEA_EXPORT_SET_INCOMPLETE")


def _validate_lifecycle(case: dict[str, object], publications: dict[str, dict[str, object]], audits: dict[str, dict[str, object]], outbox: dict[str, dict[str, object]]) -> None:
    events: dict[str, dict[str, object]] = {}
    for event in case["lifecycle_events"]:
        identity = event.get("event_id") or event.get("supersession_id") or event.get("withdrawal_id")
        _require(isinstance(identity, str) and identity, "FMEA_LIFECYCLE_BINDING_INVALID")
        _require(identity not in events, "FMEA_DUPLICATE_EVENT")
        events[identity] = event
    supersessions = [event for event in events.values() if event.get("old_publication_id") and event.get("new_publication_id")]
    withdrawals = [event for event in events.values() if event.get("publication_id") and not event.get("old_publication_id")]
    _require(len(supersessions) == 1 and len(withdrawals) == 1, "FMEA_LIFECYCLE_BINDING_INVALID")
    supersession = supersessions[0]
    withdrawal = withdrawals[0]
    publication_ids = set(publications)
    _require({supersession.get("old_publication_id"), supersession.get("new_publication_id")} <= publication_ids, "FMEA_LIFECYCLE_BINDING_INVALID")
    _require(supersession["old_publication_id"] != supersession["new_publication_id"], "FMEA_LIFECYCLE_BINDING_INVALID")
    _require(withdrawal.get("publication_id") == supersession.get("new_publication_id"), "FMEA_LIFECYCLE_BINDING_INVALID")
    for event, command, outbox_type, identity_field in (
        (supersession, "fmea.publication.supersede", "publication.superseded", "supersession_id"),
        (withdrawal, "fmea.publication.withdraw", "publication.withdrawn", "withdrawal_id"),
    ):
        identity = event.get(identity_field)
        matched_audits = [
            audit for audit in audits.values()
            if audit.get("command") == command and audit.get("row_id") == identity
        ]
        matched_outbox = [
            item for item in outbox.values()
            if item.get("event_type") == outbox_type and item.get("aggregate_id") == identity
        ]
        _require(len(matched_audits) == 1 and len(matched_outbox) == 1, "FMEA_LIFECYCLE_BINDING_INVALID")
        _require(matched_audits[0].get("actor_type") == "human", "FMEA_AUTHORITY_BINDING_INVALID")
        _require(matched_audits[0].get("actor_id") == event.get("actor_id"), "FMEA_AUTHORITY_BINDING_INVALID")
        _require(matched_outbox[0].get("workspace_id") == matched_audits[0].get("workspace_id"), "FMEA_OUTBOX_BINDING_INVALID")
        payload = matched_outbox[0].get("payload", {})
        payload_key = "supersession" if identity_field == "supersession_id" else "withdrawal"
        _require(payload.get(payload_key) == event, "FMEA_LIFECYCLE_BINDING_INVALID")
        _require(payload.get("operation") == command.removeprefix("fmea."), "FMEA_LIFECYCLE_BINDING_INVALID")
        bound_publications = {"old": supersession["old_publication_id"], "replacement": supersession["new_publication_id"]} if payload_key == "supersession" else {"publication": withdrawal["publication_id"]}
        for key, publication_id in bound_publications.items():
            _require(payload.get(key) == publications[publication_id], "FMEA_LIFECYCLE_BINDING_INVALID")
    states = _required_list(case, "publication_lifecycle")
    _require(len(states) == 2, "FMEA_LIFECYCLE_BINDING_INVALID")
    indexed = _indexed([state["publication"] for state in states], "publication_id")
    _require(indexed == publications, "FMEA_LIFECYCLE_BINDING_INVALID")
    for state in states:
        is_old = state["publication"]["publication_id"] == supersession["old_publication_id"]
        _require(state.get("effective_status") == ("superseded" if is_old else "withdrawn"), "FMEA_LIFECYCLE_BINDING_INVALID")
        _require(state.get("supersession") == (supersession if is_old else None) and state.get("withdrawal") == (None if is_old else withdrawal), "FMEA_LIFECYCLE_BINDING_INVALID")


def _validate_steps(case: dict[str, object], audits: dict[str, dict[str, object]], outbox: dict[str, dict[str, object]]) -> None:
    steps = _indexed(case["steps"], "step_id")
    commands = {step.get("command") for step in steps.values()}
    expected = {
        "evidence.select", "candidate.generate", "review.candidates.persist", "review.suggestion.start", "review.decision",
        "fmea.risk.propose", "fmea.risk.confirm", "fmea.propagation.start", "fmea.propagation.review",
        "fmea.revision.assemble", "fmea.approval.submit", "fmea.approval.decide", "fmea.publication.publish",
        "fmea.export.start", "fmea.template.import", "fmea.migration.confirm",
        "fmea.publication.supersede", "fmea.publication.withdraw",
    }
    _require(expected <= commands, "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
    for step in steps.values():
        _text(step.get("command"))
        _text(step.get("actor_id"), "FMEA_STEP_BINDING_INVALID")
        _require(step.get("actor_type") in {"system", "human", "model"}, "FMEA_STEP_BINDING_INVALID")
        _require(isinstance(step.get("before"), dict) and isinstance(step.get("after"), dict), "FMEA_STEP_BINDING_INVALID")
        _require(isinstance(step.get("result_ids"), dict), "FMEA_STEP_BINDING_INVALID")
        _require(all(isinstance(key, str) and isinstance(value, str) and value for key, value in step["result_ids"].items()), "FMEA_STEP_BINDING_INVALID")
        _require(isinstance(step.get("request_identity"), dict), "FMEA_STEP_BINDING_INVALID")
        if "request" in step:
            _require(isinstance(step["request"], dict), "FMEA_STEP_BINDING_INVALID")
        if "result" in step:
            _require(isinstance(step["result"], dict), "FMEA_STEP_BINDING_INVALID")
            for key, value in step["result_ids"].items():
                if key.endswith("_id") and key in step["result"]:
                    _require(step["result"][key] == value, "FMEA_STEP_BINDING_INVALID")
        identity = step["request_identity"]
        for key in ("request_hash", "idempotency_key_hash"):
            if key in identity:
                _digest(identity[key])
        if step.get("command") in {"fmea.approval.decide", "fmea.publication.publish", "fmea.migration.confirm", "fmea.publication.supersede", "fmea.publication.withdraw"}:
            _require(step.get("actor_type") == "human", "FMEA_AUTHORITY_BINDING_INVALID")
    _require(any(event.get("actor_type") == "human" and event.get("command") == "fmea.approval.decide" for event in audits.values()), "FMEA_AUTHORITY_BINDING_INVALID")
    _require(any(event.get("actor_type") == "human" and event.get("command") == "fmea.publication.publish" for event in audits.values()), "FMEA_AUTHORITY_BINDING_INVALID")
    _require(all(event.get("event_id") for event in audits.values()) and all(event.get("event_id") for event in outbox.values()), "FMEA_EVENT_ID_INVALID")


def _bind_replay_events(case: dict, command: str, first: dict) -> None:
    audit_id, outbox_id = first.get("audit_event_id"), first.get("outbox_event_id")
    audits = {event["event_id"]: event for event in case["audits"]}
    outbox = {event["event_id"]: event for event in case["outbox"]}
    _require(audit_id in audits and outbox_id in outbox, "FMEA_REPLAY_BINDING_INVALID")
    audit, event = audits[audit_id], outbox[outbox_id]
    _require(audit.get("command") == command and audit.get("actor_type") == "human", "FMEA_REPLAY_BINDING_INVALID")
    _require(audit.get("workspace_id") == event.get("workspace_id"), "FMEA_REPLAY_BINDING_INVALID")
    anchors = {
        "fmea.approval.decide": ("decision", "approval_id"),
        "fmea.publication.publish": ("publication", "publication_id"),
        "fmea.publication.supersede": ("supersession", "supersession_id"),
        "fmea.publication.withdraw": ("withdrawal", "withdrawal_id"),
    }
    payload_key, identity = anchors[command]
    record = event.get("payload", {}).get(payload_key, {})
    _require(bool(record) and record.get(identity) == first.get(identity) == event.get("aggregate_id") == audit.get("row_id"), "FMEA_REPLAY_BINDING_INVALID")
    for field in ("record_version", "publication_id", "snapshot_id", "manifest_id", "old_publication_id", "new_publication_id"):
        if field in first:
            _require(record.get(field) == first[field], "FMEA_REPLAY_BINDING_INVALID")


def _replay_step(case: dict, command: str, first: dict) -> str:
    """Bind each receipt to raw persisted data, then its actual command step."""
    catalog = {
        "review.decision": case["review_decisions"],
        "fmea.propagation.start": case["propagation_runs"],
        "fmea.export.start": [export["run"] for export in case["exports"]],
        "fmea.template.import": case["template_drafts"],
        "fmea.template.patch.suggest": case["template_patch_suggestions"],
        "fmea.template.patch.accept": [item["compiled"] for item in case["registered_templates"]],
        "fmea.migration.dry_run": case["migration_reports"],
        "fmea.migration.confirm": case["migration_results"],
    }
    if command in catalog:
        _require(sum(first == record for record in catalog[command]) == 1, "FMEA_REPLAY_BINDING_INVALID")
    elif command in {"fmea.risk.confirm", "fmea.propagation.review"}:
        payload_key, event_type, identity = ("assessment", "risk.confirmed", "assessment_id") if command == "fmea.risk.confirm" else ("graph", "propagation.confirmed", "graph_revision_id")
        record = first.get(payload_key, {})
        event = _one_event(case, event_type, record.get(identity))
        body = event["payload"]
        expected = {payload_key: body[payload_key], "audit_event_id": body["audit_event_id"], "decision_id": body["decision_id"], "outbox_event_id": event["event_id"], "persisted": True, "replayed": False}
        _require(first == expected, "FMEA_REPLAY_BINDING_INVALID")
    else:
        _require(command in {"fmea.approval.decide", "fmea.publication.publish", "fmea.publication.supersede", "fmea.publication.withdraw"}, "FMEA_REPLAY_BINDING_INVALID")
        _bind_replay_events(case, command, first)
    primary_ids = {
        "review.decision": ("decision_id", first.get("decision_id")),
        "fmea.risk.confirm": ("decision_id", first.get("decision_id")),
        "fmea.propagation.start": ("run_id", first.get("run_id")),
        "fmea.propagation.review": ("graph_revision_id", first.get("graph", {}).get("graph_revision_id")),
    }
    matches = []
    for step in case["steps"]:
        if step["command"] != command:
            continue
        if "result" in step:
            matched = step["result"] == first
        else:
            identity, value = primary_ids.get(command, (None, None))
            matched = value is not None and step["result_ids"].get(identity) == value
        if matched:
            matches.append(step["step_id"])
    _require(len(matches) == 1, "FMEA_REPLAY_BINDING_INVALID")
    return matches[0]


def _validate_replays(case: dict[str, object]) -> None:
    replays = _required_list(case, "replays")
    expected = {"review.decision", "fmea.risk.confirm", "fmea.propagation.start", "fmea.propagation.review", "fmea.approval.decide", "fmea.publication.publish", "fmea.publication.supersede", "fmea.publication.withdraw", "fmea.migration.confirm"}
    commands = set()
    replayed_steps = set()
    for replay in replays:
        command = _text(replay.get("command"))
        commands.add(command)
        first, retried = replay.get("first"), replay.get("replayed")
        _require(isinstance(first, dict) and bool(first) and isinstance(retried, dict), "FMEA_REPLAY_SCHEMA_INVALID")
        _require(_replay_semantics_equal(first, retried), "FMEA_REPLAY_RESULT_MISMATCH")
        if "replayed" in first or "replayed" in retried:
            _require(type(first.get("replayed")) is bool and type(retried.get("replayed")) is bool, "FMEA_REPLAY_SCHEMA_INVALID")
            _require(first["replayed"] is False and retried["replayed"] is True, "FMEA_REPLAY_SCHEMA_INVALID")
        if "same_persisted_result" in replay:
            _require(replay["same_persisted_result"] is True, "FMEA_REPLAY_RESULT_MISMATCH")
        before, after = replay.get("event_counts_before"), replay.get("event_counts_after")
        _require(isinstance(before, dict) and isinstance(after, dict) and before == after, "FMEA_REPLAY_EVENT_COUNT_MISMATCH")
        _require(all(type(value) is int and value >= 0 for value in before.values()), "FMEA_REPLAY_SCHEMA_INVALID")
        step_id = _replay_step(case, command, first)
        _require(step_id not in replayed_steps, "FMEA_DUPLICATE_REPLAY")
        replayed_steps.add(step_id)
    _require(expected <= commands, "FMEA_REPLAY_INCOMPLETE")
    required_steps = {step["step_id"] for step in case["steps"] if step["command"] in commands}
    _require(required_steps == replayed_steps, "FMEA_REPLAY_INCOMPLETE")


def _validate_used_payloads(cases: list[dict[str, object]], payloads: dict[str, bytes]) -> None:
    _require(set(payloads) == _used_payload_paths(cases), "FMEA_ARTIFACT_UNUSED_FILE")


def validate_case_semantics(
    case: dict[str, object],
    summary: dict[str, int],
    payloads: dict[str, bytes] | None = None,
    contract_version: str = SCHEMA_VERSION,
) -> None:
    """Validate one case from its raw arrays and actual export bytes."""
    _require(isinstance(case, dict), "FMEA_ARTIFACT_SCHEMA_INVALID")
    _require(case.get("case_id") == "fuel-combustion", "FMEA_CASE_INVALID")
    _require(case.get("coverage") == "full_lifecycle", "FMEA_CASE_COVERAGE_INVALID")
    for name in _FULL_REQUIRED:
        _required_list(case, name)
    _require(isinstance(case.get("analyses"), list), "FMEA_ARTIFACT_SCHEMA_INVALID")
    analysis_ids = {item.get("analysis_id") for item in case["analyses"]}
    _require(len(analysis_ids) == 1 and all(isinstance(item, str) and item for item in analysis_ids), "FMEA_SCOPE_BINDING_INVALID")
    revision_records = case["revisions"]
    workspace_ids = {item.get("workspace_id") for item in revision_records}
    _require(len(workspace_ids) == 1 and all(isinstance(item, str) and item for item in workspace_ids), "FMEA_SCOPE_BINDING_INVALID")
    workspace_id = next(iter(workspace_ids))
    analysis_id = next(iter(analysis_ids))
    _scoped(case["analyses"], workspace_id, analysis_id)
    packs = _indexed(case["evidence_packs"], "pack_id", "FMEA_DUPLICATE_EVIDENCE_PACK")
    _validate_evidence_bindings(case, packs, workspace_id, analysis_id)
    audits = _indexed(case["audits"], "event_id", "FMEA_DUPLICATE_EVENT")
    outbox = _indexed(case["outbox"], "event_id", "FMEA_DUPLICATE_EVENT")
    _validate_review_and_risk(case, packs, audits, outbox, workspace_id, analysis_id)
    revisions, root, governance_child, migration_child = _validate_revisions(case, workspace_id, analysis_id)
    _validate_graph_and_revision_links(case, packs, revisions, workspace_id, analysis_id)
    _validate_propagation_receipts(case)
    _validate_graph_structure(case, revisions)
    _validate_versioned_content(case, revisions, migration_child, contract_version)
    _validate_steps(case, audits, outbox)
    _validate_replays(case)
    publications = _indexed(case["publications"], "publication_id")
    _require(len(publications) == 2, "FMEA_PUBLICATION_BINDING_INVALID")
    approvals = _indexed(case["approvals"], "approval_id")
    _validate_submissions(case, revisions, approvals)
    verify_publication_bindings(case)
    _validate_manifests(case, revisions, publications, contract_version)
    _require(all(item.get("revision_id") in {root.get("revision_id"), governance_child.get("revision_id")} for item in publications.values()), "FMEA_PUBLICATION_BINDING_INVALID")
    _require(migration_child.get("revision_id") not in {item.get("revision_id") for item in publications.values()}, "FMEA_MIGRATION_BINDING_INVALID")
    _validate_export_wrappers(case, publications, payloads)
    if payloads is not None:
        verify_export_set(case, payloads, contract_version=contract_version)
    _validate_migration(case, root, migration_child, case["exports"], payloads)
    _validate_lifecycle(case, publications, audits, outbox)
    verify_native_hashes(case)
    counts = count_p0_violations(case)
    _require(counts == summary, "FMEA_SUMMARY_MISMATCH")
    _require(set(summary) == set(_P0_FIELDS) and all(type(summary[field]) is int and summary[field] == 0 for field in _P0_FIELDS), "FMEA_P0_VIOLATION")


def verify_native_hashes(case: dict[str, object]) -> None:
    """Recompute native content identities in addition to container byte hashes."""
    for collection, field, excludes, code in (
        ("snapshots", "snapshot_hash", {"snapshot_hash"}, "FMEA_SNAPSHOT_HASH_MISMATCH"),
        ("revisions", "revision_hash", {"revision_hash", "created_at"}, "FMEA_REVISION_HASH_MISMATCH"),
        ("migration_reports", "report_hash", {"report_hash", "created_at"}, "FMEA_MIGRATION_HASH_MISMATCH"),
    ):
        for item in case.get(collection, []):
            body = {key: value for key, value in item.items() if key not in excludes}
            _require(_hash_json(body) == _digest(item.get(field)), code)
    for event in case.get("outbox", []):
        _require(_hash_json(event["payload"]) == _digest(event["payload_hash"]), "FMEA_OUTBOX_HASH_MISMATCH")
    for pack in case.get("evidence_packs", []):
        refs = [{"evidence_id": ref["evidence_id"], "evidence_hash": ref["evidence_hash"], "locator": ref["locator"]} for ref in sorted(pack["refs"], key=lambda item: item["evidence_id"])]
        _require(len({ref["evidence_id"] for ref in refs}) == len(refs), "FMEA_DUPLICATE_EVIDENCE")
        body = refs
        if pack.get("parent_pack_refs"):
            body = {"evidence_refs": refs, "lineage": {
                "lineage_reason": pack["lineage_reason"], "lineage_schema_version": pack["lineage_schema_version"],
                "parent_pack_refs": [{"pack_id": identity, "pack_hash": digest} for identity, digest in pack["parent_pack_refs"]],
            }}
        _require(_hash_json(body, ascii_only=True) == _digest(pack["pack_hash"]), "FMEA_EVIDENCE_PACK_HASH_MISMATCH")


def _unique_records(records: object, identity: str) -> dict[str, dict]:
    _require(isinstance(records, list) and bool(records), "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
    indexed = {}
    for record in records:
        _require(isinstance(record, dict), "FMEA_ARTIFACT_SCHEMA_INVALID")
        key = record.get(identity)
        _require(isinstance(key, str) and bool(key.strip()), "FMEA_ARTIFACT_ID_INVALID")
        _require(key not in indexed, "FMEA_DUPLICATE_RECORD")
        indexed[key] = record
    return indexed


def _export_snapshot(view: dict, format_name: str) -> dict:
    media = {
        "json": "application/json",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
    _require(view.get("schema_version") == "graphrag.fmea.export.v1", "FMEA_EXPORT_SCHEMA_INVALID")
    _require(view.get("media_type") == media[format_name], "FMEA_EXPORT_FORMAT_INVALID")
    _require(view.get("draft_preview") is False and view.get("draft_marker") is None and view.get("source_publication_id") is None, "FMEA_EXPORT_NOT_PUBLISHED")
    result = {key: value for key, value in view.items() if key not in {"schema_version", "snapshot_schema_version", "format", "media_type", "draft_preview", "draft_marker", "source_publication_id"}}
    result["schema_version"] = view.get("snapshot_schema_version")
    _require(result["schema_version"] == "graphrag.fmea.normalized-snapshot.v1", "FMEA_EXPORT_SCHEMA_INVALID")
    return result


def verify_export_set(
    case: dict[str, object], payloads: dict[str, bytes], *, contract_version: str | None = None
) -> set[str]:
    """Bind independent decoded data, not just a displayed hash, to snapshots."""
    snapshots = _unique_records(case.get("snapshots"), "snapshot_id")
    if contract_version is None:
        contract_version = (
            SCHEMA_VERSION
            if any(snapshot.get("version_manifest", {}).get("body_schema_version") == BODY_SCHEMA_VERSION for snapshot in snapshots.values())
            else LEGACY_SCHEMA_VERSION
        )
    _require(contract_version in {LEGACY_SCHEMA_VERSION, SCHEMA_VERSION}, "FMEA_ARTIFACT_SCHEMA_INVALID")
    exports = _unique_records(case.get("exports"), "path")
    formats = {key: set() for key in snapshots}
    for path, record in exports.items():
        format_name = record.get("format")
        _require(format_name in {"json", "xlsx", "docx"} and path.endswith(f".{format_name}") and path in payloads, "FMEA_EXPORT_FORMAT_INVALID")
        view = parse_export(payloads[path], format_name, contract_version=contract_version)
        identity = view.get("snapshot_id")
        _require(identity in snapshots, "FMEA_EXPORT_SNAPSHOT_UNBOUND")
        _require(format_name not in formats[identity], "FMEA_DUPLICATE_EXPORT")
        _require(_export_snapshot(view, format_name) == snapshots[identity], "FMEA_EXPORT_SEMANTIC_MISMATCH")
        if contract_version == SCHEMA_VERSION and format_name in {"xlsx", "docx"}:
            _validate_visible_office_body(payloads[path], format_name, snapshots[identity])
        formats[identity].add(format_name)
    _require(all(value == {"json", "xlsx", "docx"} for value in formats.values()), "FMEA_EXPORT_SET_INCOMPLETE")
    verify_native_hashes({"snapshots": list(snapshots.values())})
    return set(exports)


def _bound_human_event(audits: dict, outbox: dict, *, command: str, resource_id: str, actor_id: str, workspace_id: str, payload_key: str, resource: dict) -> None:
    matched = [event for event in audits.values() if event.get("command") == command and event.get("row_id") == resource_id]
    _require(len(matched) == 1, "FMEA_AUDIT_BINDING_INVALID")
    audit = matched[0]
    _require(audit.get("actor_type") == "human" and audit.get("actor_id") == actor_id and audit.get("workspace_id") == workspace_id, "FMEA_AUTHORITY_BINDING_INVALID")
    stored_type = {"fmea.approval.decide": "approval.approved", "fmea.publication.publish": "publication.published"}[command]
    delivered = [event for event in outbox.values() if event.get("event_type") == stored_type and event.get("aggregate_id") == resource_id]
    _require(len(delivered) == 1, "FMEA_OUTBOX_BINDING_INVALID")
    event = delivered[0]
    _require(event.get("workspace_id") == workspace_id and event.get("payload", {}).get(payload_key) == resource, "FMEA_OUTBOX_BINDING_INVALID")
    _require(_digest(audit.get("canonical_payload_hash")) == _digest(event.get("payload_hash")), "FMEA_OUTBOX_BINDING_INVALID")


def verify_publication_bindings(case: dict[str, object]) -> None:
    """Check persisted human authorization and immutable publication identities."""
    revisions = _unique_records(case.get("revisions"), "revision_id")
    approvals = _unique_records(case.get("approvals"), "approval_id")
    publications = _unique_records(case.get("publications"), "publication_id")
    snapshots = _unique_records(case.get("snapshots"), "snapshot_id")
    audits = _unique_records(case.get("audits"), "event_id")
    outbox = _unique_records(case.get("outbox"), "event_id")
    for publication in publications.values():
        revision = revisions.get(publication.get("revision_id"), {})
        approval = approvals.get(publication.get("approval_id"), {})
        snapshot = snapshots.get(publication.get("snapshot_id"), {})
        _require(bool(revision) and bool(approval) and bool(snapshot), "FMEA_PUBLICATION_BINDING_INVALID")
        _require(approval.get("status") == "approved" and approval.get("revision_id") == revision["revision_id"], "FMEA_APPROVAL_BINDING_INVALID")
        _require(all(item.get("revision_hash") == revision["revision_hash"] for item in (approval, publication, snapshot)), "FMEA_APPROVAL_BINDING_INVALID")
        _require(snapshot.get("publication_id") == publication["publication_id"] and snapshot.get("snapshot_hash") == publication.get("snapshot_hash") and snapshot.get("manifest_id") == publication.get("manifest_id"), "FMEA_PUBLICATION_BINDING_INVALID")
        _require(all(item.get("workspace_id") == revision["workspace_id"] and item.get("analysis_id") == revision["analysis_id"] for item in (publication, snapshot)), "FMEA_PUBLICATION_SCOPE_INVALID")
        _bound_human_event(audits, outbox, command="fmea.approval.decide", resource_id=approval["approval_id"], actor_id=approval["approver_actor_id"], workspace_id=revision["workspace_id"], payload_key="decision", resource=approval)
        _bound_human_event(audits, outbox, command="fmea.publication.publish", resource_id=publication["publication_id"], actor_id=publication["publisher_actor_id"], workspace_id=revision["workspace_id"], payload_key="publication", resource=publication)
    verify_native_hashes(case)


def _validate_domain_proofs(evidence: dict[str, object]) -> None:
    proofs = evidence.get("domain_proofs", [])
    _require(isinstance(proofs, list), "FMEA_DOMAIN_PROOF_INVALID")
    _require(len(proofs) == 3 and all(isinstance(item, dict) for item in proofs), "FMEA_DOMAIN_PROOF_INVALID")
    indexed = _indexed(proofs, "pack_id", "FMEA_DUPLICATE_DOMAIN_PROOF")
    _require(set(indexed) == {"fuel-combustion", "electrical-demo", "software-demo"}, "FMEA_DOMAIN_PROOF_INVALID")
    for proof in indexed.values():
        _digest(proof.get("content_hash"))
        _require(proof.get("coverage") == "registry_compile", "FMEA_DOMAIN_PROOF_INVALID")
        _require(proof.get("kernel_schema_id") == "graphrag.fmea.v1", "FMEA_DOMAIN_PROOF_INVALID")
        for field in ("version", "template_id", "template_version", "scoring_rule_id"):
            _text(proof.get(field), "FMEA_DOMAIN_PROOF_INVALID")


def _validate_structural_case(case: dict[str, object]) -> None:
    _text(case.get("case_id"), "FMEA_CASE_INVALID")
    _require(case.get("case_id") != "fuel-combustion", "FMEA_CASE_COVERAGE_INVALID")
    _require(case.get("coverage") == "structural_domain", "FMEA_CASE_COVERAGE_INVALID")


def _validate_manifest_shape(manifest: dict[str, object], schema_version: str | None = None) -> None:
    _require(set(manifest) == {"schema_version", "artifact_id", "cases", "summary", "files"}, "FMEA_ARTIFACT_SCHEMA_INVALID")
    expected = schema_version or manifest.get("schema_version")
    _require(expected in {LEGACY_SCHEMA_VERSION, SCHEMA_VERSION} and manifest.get("schema_version") == expected, "FMEA_ARTIFACT_SCHEMA_INVALID")
    cases = manifest.get("cases")
    _require(isinstance(cases, list) and all(isinstance(case, str) and case for case in cases), "FMEA_ARTIFACT_SCHEMA_INVALID")
    _require(len(cases) == len(set(cases)), "FMEA_DUPLICATE_CASE")
    summary = manifest.get("summary")
    _require(isinstance(summary, dict) and set(summary) == set(_P0_FIELDS), "FMEA_SUMMARY_INVALID")
    _require(all(type(summary[field]) is int and summary[field] >= 0 for field in _P0_FIELDS), "FMEA_SUMMARY_INVALID")


def _used_payload_paths(cases: list[dict[str, object]]) -> set[str]:
    used = {"evidence.json"}
    for case in cases:
        if case.get("coverage") == "full_lifecycle":
            used.update(export.get("path") for export in case.get("exports", []) if isinstance(export, dict))
            used.update(source.get("path") for source in case.get("template_import_sources", []) if isinstance(source, dict))
    _require(all(isinstance(path, str) and path for path in used), "FMEA_ARTIFACT_INVENTORY_INVALID")
    return used


def verify_acceptance_directory(directory: str | Path) -> VerificationResult:
    """Verify a complete artifact while exposing only stable safe error codes."""
    artifact_id = ""
    try:
        manifest, payloads = load_bundle(directory)
        artifact_id = str(manifest.get("artifact_id", ""))
        evidence = _parse(payloads["evidence.json"])
        _require(set(evidence) >= {"schema_version", "cases"}, "FMEA_ARTIFACT_SCHEMA_INVALID")
        schema_version = manifest.get("schema_version")
        _require(evidence.get("schema_version") == schema_version, "FMEA_ARTIFACT_SCHEMA_INVALID")
        cases = evidence.get("cases")
        _require(isinstance(cases, list) and bool(cases), "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
        _require(all(isinstance(case, dict) for case in cases), "FMEA_ARTIFACT_SCHEMA_INVALID")
        case_ids = [case.get("case_id") for case in cases]
        _require(all(isinstance(case_id, str) and case_id for case_id in case_ids), "FMEA_CASE_INVALID")
        _require(len(case_ids) == len(set(case_ids)), "FMEA_DUPLICATE_CASE")
        _validate_manifest_shape(manifest, schema_version)
        manifest_cases = manifest["cases"]
        _require(set(manifest_cases) == set(case_ids), "FMEA_CASE_BINDING_INVALID")
        full_cases = [case for case in cases if case.get("case_id") == "fuel-combustion" and case.get("coverage") == "full_lifecycle"]
        _require(len(full_cases) == 1, "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
        _validate_domain_proofs(evidence)
        for case in cases:
            if case.get("coverage") == "full_lifecycle":
                validate_case_semantics(case, manifest["summary"], payloads, contract_version=schema_version)
            else:
                _validate_structural_case(case)
        _validate_used_payloads(cases, payloads)
        _require(_valid_uuid(artifact_id), "FMEA_ARTIFACT_ID_INVALID")
        root = Path(directory).absolute()
        _require(root.name == artifact_id or root.name == f".pending-{artifact_id}", "FMEA_ARTIFACT_ID_INVALID")
        _privacy(manifest)
        _privacy(evidence)
        return VerificationResult(True, artifact_id)
    except VerificationError as error:
        return VerificationResult(False, artifact_id, error.code)
    except Exception:
        return VerificationResult(False, artifact_id, "FMEA_ARTIFACT_INVALID")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("directory", nargs="?", type=Path)
    parser.add_argument("--latest", action="store_true")
    args = parser.parse_args(argv)
    if args.directory is not None and args.latest:
        parser.error("choose a directory or --latest")
    if args.directory is not None:
        result = verify_acceptance_directory(args.directory)
    elif args.latest:
        try:
            result = verify_acceptance_directory(resolve_latest_directory())
        except VerificationError as error:
            result = VerificationResult(False, "", error.code)
    else:
        result = VerificationResult(False, "", "FMEA_WORKFLOW_EVIDENCE_INCOMPLETE")
    print(json.dumps({"passed": result.passed, "artifact_id": result.artifact_id, "error_code": result.error_code}, sort_keys=True))
    return 0 if result.passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
