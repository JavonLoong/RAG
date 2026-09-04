from __future__ import annotations

import io
import json
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import orjson
import pytest
from defusedxml.ElementTree import iterparse

from core_domain.fmea.governance import canonical_json_value
from fmea_application.snapshot_contracts import (
    NormalizedFmeaSnapshot,
    iter_normalized_snapshot_pages,
)
from fmea_infrastructure.export_docx import DocxFmeaExporter
from fmea_infrastructure.export_json import CanonicalJsonExporter
from fmea_infrastructure.export_xlsx import XlsxFmeaExporter
from tests.fmea_governance_fixtures import make_normalized_snapshot

ROW_COUNT = 10_000
API_PAGE_SIZE = 100
JSON_CHUNK_SIZE = 64 * 1024
FORMAT_MEDIA_TYPES = {
    ".json": "application/json",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}
TRANSPORT_FIELDS = {"format", "media_type"}
SNAPSHOT_FIELDS = (
    "snapshot_schema_version",
    "snapshot_id",
    "workspace_id",
    "analysis_id",
    "revision_id",
    "revision_hash",
    "publication_id",
    "source_publication_id",
    "manifest_id",
    "rows",
    "risk_records",
    "propagation",
    "evidence_summary",
    "decision_summary",
    "version_manifest",
    "unresolved_items",
    "audit_summary",
    "row_count",
    "snapshot_hash",
    "created_at",
)
TABLE_FIELDS = (
    "rows",
    "risk_records",
    "propagation",
    "evidence_summary",
    "decision_summary",
    "unresolved_items",
)
MANIFEST_SNAPSHOT_FIELDS = (
    "snapshot_schema_version",
    "snapshot_id",
    "workspace_id",
    "analysis_id",
    "revision_id",
    "revision_hash",
    "publication_id",
    "source_publication_id",
    "manifest_id",
    "row_count",
    "snapshot_hash",
    "created_at",
)


@dataclass(frozen=True)
class ScaleExportResult:
    row_count: int
    output_extensions: tuple[str, ...]
    decoded_row_counts: tuple[tuple[str, int], ...]
    snapshot_hashes: tuple[tuple[str, str], ...]
    page_count: int
    observed_page_sizes: tuple[int, ...]
    max_api_page_rows: int
    json_chunk_count: int
    max_json_chunk_bytes: int
    payload_sizes: tuple[tuple[str, int], ...]
    format_elapsed_seconds: tuple[tuple[str, float], ...]
    elapsed_seconds: float

    def metadata(self) -> dict[str, Any]:
        return asdict(self)


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _decode_cell(value: object, value_type: str) -> object:
    if value_type == "str":
        return "" if value is None else str(value)
    if value_type == "null":
        return None
    if value_type == "bool":
        return json.loads(str(value).lower())
    if value_type in {"int", "float", "json"}:
        return json.loads(str(value))
    raise AssertionError


def _parse_typed_table(rows: list[tuple[object, ...]]) -> list[dict[str, object]]:
    headers = [str(value) for value in rows[0]]
    type_column = headers.index("__types__")
    decoded: list[dict[str, object]] = []
    for row in rows[1:]:
        if not any(value is not None and value != "" for value in row):
            continue
        types = json.loads(str(row[type_column]))
        identity_column = headers.index("Identity")
        item: dict[str, object] = {"Identity": _decode_cell(row[identity_column], "str")}
        for header, value in zip(headers, row, strict=True):
            if header in {"Identity", "__types__"} or header not in types:
                continue
            item[header] = _decode_cell(value, types[header])
        decoded.append(item)
    return decoded


def _parse_xlsx(payload: bytes) -> dict[str, object]:
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False, read_only=True)
    try:
        manifest_rows = list(workbook["Manifest"].iter_rows(values_only=True))
        manifest = {
            str(row[0]): _decode_cell(row[1], str(row[2]))
            for row in manifest_rows[1:]
            if row[0] is not None
        }
        result: dict[str, object] = {field: manifest[field] for field in MANIFEST_SNAPSHOT_FIELDS}
        result.update(
            {
                "schema_version": manifest["schema_version"],
                "rows": _parse_typed_table(list(workbook["FMEA"].iter_rows(values_only=True))),
                "risk_records": _parse_typed_table(list(workbook["Risk"].iter_rows(values_only=True))),
                "evidence_summary": _parse_typed_table(list(workbook["Evidence"].iter_rows(values_only=True))),
                "decision_summary": _parse_typed_table(list(workbook["Decisions"].iter_rows(values_only=True))),
                "unresolved_items": _parse_typed_table(list(workbook["Unresolved"].iter_rows(values_only=True))),
                "version_manifest": manifest["version_manifest"],
                "audit_summary": manifest["audit_summary"],
                "draft_preview": manifest["draft_preview"],
                "draft_marker": manifest["draft_marker"],
                "format": manifest["format"],
                "media_type": manifest["media_type"],
            }
        )
        propagation_rows = _parse_typed_table(list(workbook["Propagation"].iter_rows(values_only=True)))
        result["propagation"] = propagation_rows[0] if propagation_rows else None
        return result
    finally:
        workbook.close()


def _tamper_xlsx_fmea_identity(payload: bytes) -> bytes:
    output = io.BytesIO()
    with ZipFile(io.BytesIO(payload)) as source, ZipFile(output, "w", compression=ZIP_DEFLATED) as target:
        for name in source.namelist():
            part = source.read(name)
            if name == "xl/worksheets/sheet2.xml":
                part = part.replace(b">row-0<", b">tampered-row-0<", 1)
            target.writestr(name, part)
    return output.getvalue()


def _parse_docx_tables(payload: bytes) -> list[list[tuple[str, ...]]]:
    with ZipFile(io.BytesIO(payload)) as archive:
        xml_payload = archive.read("word/document.xml")

    tables: list[list[tuple[str, ...]]] = []
    current_table: list[tuple[str, ...]] | None = None
    current_row: list[str] | None = None
    current_cell: list[str] | None = None
    for event, element in iterparse(io.BytesIO(xml_payload), events=("start", "end")):
        local_name = _local_name(element.tag)
        if event == "start":
            if local_name == "tbl":
                current_table = []
            elif local_name == "tr":
                current_row = []
            elif local_name == "tc":
                current_cell = []
            continue

        if local_name == "t" and current_cell is not None:
            current_cell.append(element.text or "")
            element.clear()
        elif local_name == "tc" and current_row is not None:
            current_row.append("".join(current_cell or ()))
            current_cell = None
            element.clear()
        elif local_name == "tr" and current_table is not None:
            current_table.append(tuple(current_row or ()))
            current_row = None
            element.clear()
        elif local_name == "tbl":
            tables.append(current_table or [])
            current_table = None
            element.clear()
    return tables


def _parse_docx(payload: bytes) -> dict[str, object]:
    tables = _parse_docx_tables(payload)
    assert len(tables) >= 7
    manifest = {
        row[0]: _decode_cell(row[1], row[2])
        for row in tables[0][1:]
        if len(row) >= 3
    }
    result: dict[str, object] = {field: manifest[field] for field in MANIFEST_SNAPSHOT_FIELDS}
    result.update(
        {
            "schema_version": manifest["schema_version"],
            "rows": _parse_typed_table(tables[1]),
            "risk_records": _parse_typed_table(tables[2]),
            "evidence_summary": _parse_typed_table(tables[4]),
            "decision_summary": _parse_typed_table(tables[5]),
            "unresolved_items": _parse_typed_table(tables[6]),
            "version_manifest": manifest["version_manifest"],
            "audit_summary": manifest["audit_summary"],
            "draft_preview": manifest["draft_preview"],
            "draft_marker": manifest["draft_marker"],
            "format": manifest["format"],
            "media_type": manifest["media_type"],
        }
    )
    propagation_rows = _parse_typed_table(tables[3])
    result["propagation"] = propagation_rows[0] if propagation_rows else None
    return result


def _snapshot_identity(snapshot: NormalizedFmeaSnapshot) -> dict[str, object]:
    values = {
        "snapshot_schema_version": snapshot.schema_version,
        "snapshot_id": snapshot.snapshot_id,
        "workspace_id": snapshot.workspace_id,
        "analysis_id": snapshot.analysis_id,
        "revision_id": snapshot.revision_id,
        "revision_hash": snapshot.revision_hash,
        "publication_id": snapshot.publication_id,
        "source_publication_id": None,
        "manifest_id": snapshot.manifest_id,
        "rows": snapshot.rows,
        "risk_records": snapshot.risk_records,
        "propagation": snapshot.propagation,
        "evidence_summary": snapshot.evidence_summary,
        "decision_summary": snapshot.decision_summary,
        "version_manifest": snapshot.version_manifest,
        "unresolved_items": snapshot.unresolved_items,
        "audit_summary": snapshot.audit_summary,
        "row_count": snapshot.row_count,
        "snapshot_hash": snapshot.snapshot_hash,
        "created_at": snapshot.created_at,
    }
    return canonical_json_value(values)  # type: ignore[return-value]


def _without_transport(view: dict[str, object]) -> dict[str, object]:
    return {key: value for key, value in view.items() if key not in TRANSPORT_FIELDS}


def _without_display_identities(view: dict[str, object]) -> dict[str, object]:
    normalized = dict(view)
    for field in TABLE_FIELDS:
        records = view[field]
        if field == "propagation" and isinstance(records, dict):
            normalized[field] = {key: value for key, value in records.items() if key != "Identity"}
        elif isinstance(records, list):
            normalized[field] = [
                {key: value for key, value in record.items() if key != "Identity"}
                for record in records
            ]
    return normalized


def _record_identity(record: Any, identity_field: str | None, index: int) -> str:
    if identity_field is not None:
        value = record.get(identity_field)
        assert isinstance(value, str) and value.strip()
        return value
    for field_name in ("source_id", "code", "item_id"):
        value = record.get(field_name)
        if isinstance(value, str) and value.strip():
            return value
    return f"item-{index:03d}"


def _expected_table_identities(snapshot: NormalizedFmeaSnapshot) -> dict[str, tuple[str, ...]]:
    table_records: dict[str, tuple[Any, ...]] = {
        "rows": tuple(snapshot.rows),
        "risk_records": tuple(snapshot.risk_records),
        "propagation": () if snapshot.propagation is None else (snapshot.propagation,),
        "evidence_summary": tuple(snapshot.evidence_summary),
        "decision_summary": tuple(snapshot.decision_summary),
        "unresolved_items": tuple(snapshot.unresolved_items),
    }
    identity_fields = {
        "rows": "row_id",
        "risk_records": "assessment_id",
        "propagation": None,
        "evidence_summary": "pack_id",
        "decision_summary": "decision_id",
        "unresolved_items": None,
    }
    return {
        field: tuple(_record_identity(record, identity_fields[field], index) for index, record in enumerate(records, 1))
        for field, records in table_records.items()
    }


def _assert_table_identity_equality(
    snapshot: NormalizedFmeaSnapshot,
    views: dict[str, dict[str, object]],
) -> None:
    expected = _expected_table_identities(snapshot)
    for extension, view in views.items():
        if extension == ".json":
            continue
        for field, expected_identities in expected.items():
            records = view[field]
            if field == "propagation":
                records = () if records is None else (records,)
            actual_identities = tuple(record.get("Identity") for record in records)
            assert actual_identities == expected_identities
            assert len(actual_identities) == len(set(actual_identities))


def _assert_identity(snapshot: NormalizedFmeaSnapshot, views: dict[str, dict[str, object]]) -> None:
    expected_identity = _snapshot_identity(snapshot)
    expected_view = views[".json"]
    assert {key: expected_view[key] for key in SNAPSHOT_FIELDS} == expected_identity
    for extension, view in views.items():
        normalized_view = _without_display_identities(view)
        assert _without_transport(normalized_view) == _without_transport(expected_view)
        assert {key: normalized_view[key] for key in SNAPSHOT_FIELDS} == expected_identity
        assert view["format"] == extension.removeprefix(".")
        assert view["media_type"] == FORMAT_MEDIA_TYPES[extension]
    _assert_table_identity_equality(snapshot, views)


def _write_json_chunks(path: Path, chunks) -> tuple[int, int]:
    chunk_count = 0
    max_chunk_bytes = 0
    with path.open("wb") as output:
        for chunk in chunks:
            assert type(chunk) is bytes and 0 < len(chunk) <= JSON_CHUNK_SIZE
            output.write(chunk)
            chunk_count += 1
            max_chunk_bytes = max(max_chunk_bytes, len(chunk))
    return chunk_count, max_chunk_bytes


def _observe_pages(snapshot: NormalizedFmeaSnapshot) -> tuple[int, tuple[int, ...], int]:
    pages = list(iter_normalized_snapshot_pages(snapshot, page_size=API_PAGE_SIZE))
    assert pages
    assert tuple(row for page in pages for row in page.rows) == snapshot.rows
    observed_sizes = tuple(len(page.rows) for page in pages)
    assert max(observed_sizes) <= API_PAGE_SIZE
    for index, page in enumerate(pages):
        expected_next_offset = (index + 1) * API_PAGE_SIZE
        if expected_next_offset >= snapshot.row_count:
            expected_next_offset = None
        assert page.next_offset == expected_next_offset
    return len(pages), tuple(sorted(set(observed_sizes))), max(observed_sizes)


def _export_scale_fixture(snapshot: NormalizedFmeaSnapshot, output_dir: Path) -> ScaleExportResult:
    page_count, observed_page_sizes, max_api_page_rows = _observe_pages(snapshot)
    started = time.perf_counter()
    exporters = (
        (".json", CanonicalJsonExporter()),
        (".xlsx", XlsxFmeaExporter()),
        (".docx", DocxFmeaExporter()),
    )
    format_elapsed: list[tuple[str, float]] = []
    json_chunk_count = 0
    max_json_chunk_bytes = 0
    payload_sizes: list[tuple[str, int]] = []
    for extension, exporter in exporters:
        format_started = time.perf_counter()
        path = output_dir / f"fmea-scale{extension}"
        if extension == ".json":
            json_chunk_count, max_json_chunk_bytes = _write_json_chunks(
                path, exporter.iter_chunks(snapshot, chunk_size=JSON_CHUNK_SIZE),
            )
        else:
            path.write_bytes(exporter.render(snapshot))
        format_elapsed.append((extension, time.perf_counter() - format_started))
        payload_sizes.append((extension, path.stat().st_size))

    payloads = {
        extension: (output_dir / f"fmea-scale{extension}").read_bytes()
        for extension, _ in exporters
    }
    views = {
        ".json": orjson.loads(payloads[".json"]),
        ".xlsx": _parse_xlsx(payloads[".xlsx"]),
        ".docx": _parse_docx(payloads[".docx"]),
    }
    _assert_identity(snapshot, views)
    return ScaleExportResult(
        row_count=snapshot.row_count,
        output_extensions=tuple(sorted(payloads)),
        decoded_row_counts=tuple((extension, len(view["rows"])) for extension, view in views.items()),
        snapshot_hashes=tuple((extension, str(view["snapshot_hash"])) for extension, view in views.items()),
        page_count=page_count,
        observed_page_sizes=observed_page_sizes,
        max_api_page_rows=max_api_page_rows,
        json_chunk_count=json_chunk_count,
        max_json_chunk_bytes=max_json_chunk_bytes,
        payload_sizes=tuple(payload_sizes),
        format_elapsed_seconds=tuple(format_elapsed),
        elapsed_seconds=time.perf_counter() - started,
    )


@pytest.fixture
def export_large_fixture(tmp_path: Path):
    def export(*, row_count: int) -> ScaleExportResult:
        snapshot = make_normalized_snapshot(rows=row_count)
        return _export_scale_fixture(snapshot, tmp_path)

    return export


def test_small_export_decodes_three_rows_through_all_existing_exporters(tmp_path: Path) -> None:
    snapshot = make_normalized_snapshot(rows=3)
    pages = list(iter_normalized_snapshot_pages(snapshot, page_size=2))
    assert [len(page.rows) for page in pages] == [2, 1]
    assert pages[-1].next_offset is None

    payloads = {
        ".json": CanonicalJsonExporter().render(snapshot),
        ".xlsx": XlsxFmeaExporter().render(snapshot),
        ".docx": DocxFmeaExporter().render(snapshot),
    }
    for extension, payload in payloads.items():
        (tmp_path / f"small-export{extension}").write_bytes(payload)

    views = {
        ".json": orjson.loads(payloads[".json"]),
        ".xlsx": _parse_xlsx(payloads[".xlsx"]),
        ".docx": _parse_docx(payloads[".docx"]),
    }
    _assert_identity(snapshot, views)
    assert {path.suffix for path in tmp_path.iterdir()} == {".json", ".xlsx", ".docx"}
    assert {len(view["rows"]) for view in views.values()} == {snapshot.row_count}


def test_tampered_xlsx_display_identity_is_observable_and_rejected() -> None:
    snapshot = make_normalized_snapshot(rows=1)
    payload = XlsxFmeaExporter().render(snapshot)

    tampered_view = _parse_xlsx(_tamper_xlsx_fmea_identity(payload))

    assert tampered_view["rows"][0].get("Identity") == "tampered-row-0"
    with pytest.raises(AssertionError):
        _assert_table_identity_equality(snapshot, {".xlsx": tampered_view})


def test_10000_row_export_preserves_identity_with_bounded_pages_and_json_chunks(export_large_fixture) -> None:
    result = export_large_fixture(row_count=ROW_COUNT)

    assert result.row_count == ROW_COUNT
    assert set(result.output_extensions) == set(FORMAT_MEDIA_TYPES)
    assert all(decoded_count == result.row_count for _, decoded_count in result.decoded_row_counts)
    assert len(result.decoded_row_counts) == len(result.output_extensions)
    assert len({snapshot_hash for _, snapshot_hash in result.snapshot_hashes}) == 1
    assert result.max_api_page_rows <= API_PAGE_SIZE
    assert result.page_count >= 1
    assert result.observed_page_sizes
    assert max(result.observed_page_sizes) == result.max_api_page_rows
    assert result.json_chunk_count >= 1
    assert result.max_json_chunk_bytes <= JSON_CHUNK_SIZE
    print(json.dumps(result.metadata(), sort_keys=True))
