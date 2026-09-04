from __future__ import annotations

# ruff: noqa: RUF001
import io
from dataclasses import fields
from zipfile import ZipFile

import openpyxl
import pytest
from defusedxml.ElementTree import fromstring as safe_xml_fromstring

from fmea_application.snapshot_contracts import NormalizedFmeaSnapshot, snapshot_content_hash
from fmea_infrastructure.export_xlsx import XlsxFmeaExporter
from tests.fmea_governance_fixtures import make_normalized_snapshot

MARKER = "DRAFT PREVIEW — NOT PUBLISHED"
SHEETS = ["正文", "正文详情", "Manifest", "FMEA", "Risk", "Propagation", "Evidence", "Decisions", "Unresolved"]


def _forge_snapshot(snapshot: NormalizedFmeaSnapshot, **overrides: object) -> NormalizedFmeaSnapshot:
    forged = object.__new__(NormalizedFmeaSnapshot)
    for field in fields(snapshot):
        object.__setattr__(forged, field.name, getattr(snapshot, field.name))
    for name, value in overrides.items():
        object.__setattr__(forged, name, value)
    return forged


def test_xlsx_render_has_the_exact_readable_sheet_contract() -> None:
    snapshot = make_normalized_snapshot(rows=2)
    exporter = XlsxFmeaExporter()

    payload = exporter.render(snapshot)

    assert type(payload) is bytes
    assert exporter.format == "xlsx"
    assert exporter.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    assert workbook.sheetnames == SHEETS
    assert workbook["Manifest"].freeze_panes == "A2"
    assert workbook["FMEA"].freeze_panes == "A2"
    manifest = {
        str(row[0].value): row[1].value for row in workbook["Manifest"].iter_rows(min_row=2) if row[0].value is not None
    }
    assert manifest["schema_version"] == "graphrag.fmea.export.v1"
    assert manifest["snapshot_id"] == snapshot.snapshot_id
    assert manifest["revision_id"] == snapshot.revision_id
    assert manifest["publication_id"] == snapshot.publication_id
    assert manifest["manifest_id"] == snapshot.manifest_id
    assert manifest["snapshot_hash"] == snapshot.snapshot_hash
    assert manifest["row_count"] == "2"
    assert manifest["version_manifest"]
    assert manifest["draft_marker"] == "null"
    for worksheet in workbook.worksheets:
        assert worksheet.auto_filter.ref is not None
        assert all(0 < dimension.width <= 48 for dimension in worksheet.column_dimensions.values())


def test_xlsx_reading_body_uses_report_view_and_keeps_full_evidence_in_deduplicated_details() -> None:
    from tests.unit.test_fmea_report_view import _layout, _snapshot

    quote = "滤清器堵塞会造成燃料压力下降，维护记录应核对更换周期。" * 10
    snapshot = _snapshot(
        layout=_layout(
            {
                "field_key": "failure_mode",
                "label": "故障模式",
                "value_type": "string",
                "value_path": ("row", "failure_mode"),
            },
            {
                "field_key": "causes",
                "label": "原因",
                "value_type": "string[]",
                "value_path": ("row", "causes"),
            },
            {
                "field_key": "effects",
                "label": "影响",
                "value_type": "string[]",
                "value_path": ("row", "effects"),
            },
            {
                "field_key": "fuel.pressure_drop",
                "label": "压降",
                "value_type": "decimal",
                "value_path": ("extension_values", "fuel.pressure_drop"),
            },
        ),
        row={
            "failure_mode": "燃料滤清器堵塞",
            "causes": ("杂质积聚", "低温结蜡", "维护周期过长"),
            "effects": ("供油压力下降", "燃烧不稳定"),
            "extension_values": (
                {"field_key": "fuel.pressure_drop", "value_type": "decimal", "value": "48.2000"},
            ),
        },
        refs=(
            {
                "evidence_id": "evidence-1",
                "document_id": "manual-1",
                "document_version": "7",
                "content_hash": "c" * 64,
                "evidence_hash": "d" * 64,
                "locator": {"page": 12, "span": 3},
                "quote": quote,
                "source_type": "primary_document",
                "source_trust": "trusted",
            },
        ),
    )

    workbook = openpyxl.load_workbook(io.BytesIO(XlsxFmeaExporter().render(snapshot)), data_only=False)
    body_rows = list(workbook["正文"].iter_rows(values_only=True))
    detail_rows = list(workbook["正文详情"].iter_rows(values_only=True))
    body_headers = set(body_rows[0])
    body_values = {value for row in body_rows[1:] for value in row}
    detail_values = {value for row in detail_rows[1:] for value in row}

    assert {"故障模式", "原因", "影响", "压降"} <= body_headers
    assert "燃料滤清器堵塞" in body_values
    assert "row-1" not in body_values
    assert {"row-1", "fuel.pressure_drop", "48.2000"} <= detail_values
    assert any(quote in value for value in detail_values if isinstance(value, str))
    assert any(row[3] == "evidence-1.quote" and row[4] == quote for row in detail_rows[1:])
    assert any(row[3] == "evidence-1.document_id" and row[4] == "manual-1" for row in detail_rows[1:])
    assert not any('"quote":' in str(row[4]) for row in detail_rows[1:])
    assert detail_rows[0] == ("行ID", "记录版本", "详情类型", "字段", "内容")
    assert workbook["正文详情"].cell(2, 5).alignment.wrap_text is True
    assert workbook["正文详情"].row_dimensions[2].height >= 36
    quote_row_numbers = [
        row_number
        for row_number, row in enumerate(detail_rows[1:], start=2)
        if row[3] == "evidence-1.quote"
    ]
    assert len(quote_row_numbers) == 1
    assert workbook["正文详情"].row_dimensions[quote_row_numbers[0]].height >= 150


def test_xlsx_human_quote_continuation_reassembles_literal_text() -> None:
    from tests.unit.test_fmea_report_view import _layout, _snapshot

    long_quote = "可追溯证据文本。" * 100
    snapshot = _snapshot(
        layout=_layout({
            "field_key": "failure_mode",
            "label": "故障模式",
            "value_type": "string",
            "value_path": ("row", "failure_mode"),
        }),
        row={"failure_mode": "燃料压力下降"},
        refs=(
            {
                "evidence_id": "evidence-long",
                "document_id": "manual-long",
                "document_version": "1",
                "content_hash": "c" * 64,
                "evidence_hash": "d" * 64,
                "locator": {"page": 1, "span": 1},
                "quote": long_quote,
                "source_type": "primary_document",
                "source_trust": "trusted",
            },
        ),
    )

    workbook = openpyxl.load_workbook(io.BytesIO(XlsxFmeaExporter().render(snapshot)), data_only=False)
    detail_rows = list(workbook["正文详情"].iter_rows(values_only=True))[1:]
    quote_rows = [row for row in detail_rows if str(row[3]).startswith("evidence-long.quote [part ")]

    assert [row[3] for row in quote_rows] == [
        "evidence-long.quote [part 1/2]",
        "evidence-long.quote [part 2/2]",
    ]
    assert "".join(str(row[4]) for row in quote_rows) == long_quote


def test_xlsx_human_continuations_cover_newlines_main_and_ordinary_details_losslessly() -> None:
    from tests.unit.test_fmea_report_view import _layout, _snapshot

    newline_text = "\n".join("甲甲" for _ in range(100))
    main_long_text = "主正文" * 250
    detail_long_text = "普通详情" * 250
    snapshot = _snapshot(
        layout=_layout(
            {
                "field_key": "failure_mode",
                "label": "故障模式",
                "value_type": "string",
                "value_path": ("row", "failure_mode"),
            },
            {
                "field_key": "item_id",
                "label": "项目",
                "value_type": "string",
                "value_path": ("row", "item_id"),
            },
        ),
        row={
            "failure_mode": newline_text,
            "item_id": main_long_text,
            "ordinary_detail": detail_long_text,
        },
        refs=(
            {
                "evidence_id": "evidence-empty-locator",
                "document_id": "manual-empty-locator",
                "document_version": "1",
                "content_hash": "c" * 64,
                "evidence_hash": "d" * 64,
                "locator": {},
                "quote": "短引文",
                "source_type": "primary_document",
                "source_trust": "trusted",
            },
        ),
    )

    workbook = openpyxl.load_workbook(io.BytesIO(XlsxFmeaExporter().render(snapshot)), data_only=False)
    body_rows = list(workbook["正文"].iter_rows(values_only=True))
    body_headers = list(body_rows[0])
    continuation_column = body_headers.index("续字段")

    for field_key, label, expected in (
        ("failure_mode", "故障模式", newline_text),
        ("item_id", "项目", main_long_text),
    ):
        value_column = body_headers.index(label)
        parts = [
            row
            for row in body_rows[1:]
            if field_key in str(row[continuation_column])
        ]
        assert len(parts) > 1
        assert "".join(str(row[value_column] or "") for row in parts) == expected

    detail_rows = list(workbook["正文详情"].iter_rows(values_only=True))[1:]
    detail_parts = [row for row in detail_rows if str(row[3]).startswith("ordinary_detail [part ")]
    assert len(detail_parts) > 1
    assert "".join(str(row[4]) for row in detail_parts) == detail_long_text
    assert any(row[3] == "evidence-empty-locator.locator" and row[4] == "（空对象）" for row in detail_rows)


def test_xlsx_reading_body_keeps_all_saved_view_columns_in_declared_order() -> None:
    from tests.unit.test_fmea_report_view import _layout, _snapshot

    field_specs = (
        ("item_id", "项目"),
        ("function_id", "功能"),
        ("failure_mode", "故障模式"),
        ("causes", "原因"),
        ("effects", "影响"),
        ("controls", "控制"),
        ("barriers", "屏障"),
        ("actions", "措施"),
        ("custom.control", "自定义控制"),
    )
    columns = tuple(
        {
            "field_key": field_key,
            "label": label,
            "value_type": "string",
            "value_path": ("extension_values", field_key)
            if field_key == "custom.control"
            else ("row", field_key),
        }
        for field_key, label in field_specs
    )
    snapshot = _snapshot(
        layout=_layout(*columns),
        row={
            "item_id": "item-1",
            "function_id": "function-1",
            "failure_mode": "低压",
            "causes": ("堵塞",),
            "effects": ("供油不稳",),
            "controls": ("压差监测",),
            "barriers": ("报警",),
            "actions": ("更换滤芯",),
            "extension_values": (
                {"field_key": "custom.control", "value_type": "string", "value": "自定义联锁"},
            ),
        },
    )

    workbook = openpyxl.load_workbook(io.BytesIO(XlsxFmeaExporter().render(snapshot)), data_only=False)
    body_rows = list(workbook["正文"].iter_rows(values_only=True))
    headers = list(body_rows[0])
    expected_labels = [label for _, label in field_specs]
    assert [header for header in headers if header in expected_labels] == expected_labels
    body_values = {value for row in body_rows[1:] for value in row}
    assert {"更换滤芯", "自定义联锁"} <= body_values


def test_xlsx_rejects_reading_detail_that_exceeds_excel_cell_limit_without_truncation() -> None:
    from tests.unit.test_fmea_report_view import _layout, _snapshot

    snapshot = _snapshot(
        layout=_layout({
            "field_key": "failure_mode",
            "label": "故障模式",
            "value_type": "string",
            "value_path": ("row", "failure_mode"),
        }),
        row={"failure_mode": "长正文" * 20_000},
    )

    with pytest.raises(ValueError) as captured:
        XlsxFmeaExporter().render(snapshot)

    assert captured.value.code == "FMEA_EXPORT_XLSX_INVALID"
    assert str(captured.value) == "snapshot cannot be rendered as XLSX"


def test_xlsx_formula_prefixes_are_real_strings_and_never_formula_xml() -> None:
    snapshot = make_normalized_snapshot(
        row_payload={
            "row_id": "row-injection",
            "formula": "=SUM(A1:A2)",
            "plus": "+cmd|' /C calc'!A0",
            "minus": "-1+1",
            "at": "@SUM(1,1)",
        }
    )

    payload = XlsxFmeaExporter().render(snapshot)

    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    worksheet = workbook["FMEA"]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for field_name, expected in {
        "formula": "=SUM(A1:A2)",
        "plus": "+cmd|' /C calc'!A0",
        "minus": "-1+1",
        "at": "@SUM(1,1)",
    }.items():
        cell = worksheet.cell(2, headers[field_name])
        assert cell.value == expected
        assert cell.data_type == "s"

    with ZipFile(io.BytesIO(payload)) as archive:
        original_names = archive.namelist()
        names = {name.casefold() for name in original_names}
        assert not any(name.startswith("xl/externallinks/") for name in names)
        assert not any(name.endswith((".xlsm", ".bin")) for name in names)
        for name in original_names:
            if name.endswith(".xml"):
                root = safe_xml_fromstring(archive.read(name))
                assert not list(root.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f"))
        assert b"<f>" not in archive.read("xl/worksheets/sheet2.xml")


def test_xlsx_draft_marker_is_explicit_and_published_output_has_none() -> None:
    snapshot = make_normalized_snapshot()
    exporter = XlsxFmeaExporter(draft_preview=True)

    draft = exporter.render(snapshot)
    published = exporter.render(snapshot, draft_preview=False)

    draft_workbook = openpyxl.load_workbook(io.BytesIO(draft), read_only=False)
    published_workbook = openpyxl.load_workbook(io.BytesIO(published), read_only=False)
    draft_text = [cell.value for row in draft_workbook["Manifest"].iter_rows() for cell in row]
    published_text = [cell.value for row in published_workbook["Manifest"].iter_rows() for cell in row]
    assert MARKER in draft_text
    assert MARKER not in published_text
    draft_manifest = {
        str(row[0].value): (row[1].value, row[2].value)
        for row in draft_workbook["Manifest"].iter_rows(min_row=2)
        if row[0].value is not None
    }
    published_manifest = {
        str(row[0].value): (row[1].value, row[2].value)
        for row in published_workbook["Manifest"].iter_rows(min_row=2)
        if row[0].value is not None
    }
    assert draft_manifest["publication_id"] == ("null", "null")
    assert draft_manifest["source_publication_id"] == (snapshot.publication_id, "str")
    assert published_manifest["publication_id"] == (snapshot.publication_id, "str")
    assert published_manifest["source_publication_id"] == ("null", "null")


@pytest.mark.parametrize("bad_value", ["bad\x01value", float("nan"), float("inf"), object()])
def test_xlsx_rejects_office_unsafe_or_malformed_snapshot_without_leaking_input(bad_value: object) -> None:
    snapshot = make_normalized_snapshot()
    forged = _forge_snapshot(snapshot, rows=({"row_id": "row-1", "value": bad_value},), row_count=1)

    with pytest.raises(ValueError) as captured:
        XlsxFmeaExporter().render(forged)

    assert getattr(captured.value, "code", None) == "FMEA_EXPORT_XLSX_INVALID"
    assert str(captured.value) == "snapshot cannot be rendered as XLSX"
    assert "bad" not in str(captured.value)


@pytest.mark.parametrize(
    "overrides",
    (
        {"row_count": 0},
        {"schema_version": "not-normalized-v99"},
    ),
)
def test_xlsx_rejects_hash_consistent_snapshot_invariant_bypass(overrides: dict[str, object]) -> None:
    snapshot = make_normalized_snapshot()
    forged = _forge_snapshot(snapshot, **overrides)
    object.__setattr__(forged, "snapshot_hash", snapshot_content_hash(forged))

    with pytest.raises(ValueError) as captured:
        XlsxFmeaExporter().render(forged)

    assert getattr(captured.value, "code", None) == "FMEA_EXPORT_XLSX_INVALID"
    assert str(captured.value) == "snapshot cannot be rendered as XLSX"
    assert captured.value.__cause__ is None


def test_xlsx_rejects_wrong_type_and_does_not_mutate_snapshot() -> None:
    snapshot = make_normalized_snapshot()
    before = repr(snapshot)

    with pytest.raises(ValueError) as captured:
        XlsxFmeaExporter().render(object())  # type: ignore[arg-type]

    assert getattr(captured.value, "code", None) == "FMEA_EXPORT_SNAPSHOT_INVALID"
    assert str(captured.value) == "snapshot must be a NormalizedFmeaSnapshot"
    XlsxFmeaExporter().render(snapshot)
    assert repr(snapshot) == before
