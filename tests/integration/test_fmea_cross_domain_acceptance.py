"""Cross-domain portability fixtures for the FMEA product contracts."""

from __future__ import annotations

import ast
import io
import json
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import orjson
import pytest
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph

from core_domain.fmea.codec import decode_row, encode_json
from core_domain.fmea.domain_pack import DomainPackManifest
from core_domain.fmea.entities import FieldClaim, FieldValue, FmeaRow, validate_extension_values
from core_domain.fmea.states import ClaimStatus, EvidenceSupportStatus, PublicationStatus, ReviewStatus
from fmea_application.snapshot_contracts import NormalizedFmeaSnapshot
from fmea_infrastructure.domain_pack_registry import (
    FileDomainPackRegistry,
    FileScoringRuleRegistry,
    load_domain_pack_manifest,
    load_scoring_rule_pack,
)
from fmea_infrastructure.export_docx import DocxFmeaExporter
from fmea_infrastructure.export_json import CanonicalJsonExporter
from fmea_infrastructure.export_xlsx import XlsxFmeaExporter
from structured_output_application import TemplateCompiler
from structured_output_infrastructure import Draft202012SchemaAdapter, FileTemplateRegistry, load_template_source
from tests.fmea_governance_fixtures import make_fmea_revision, make_normalized_snapshot

REPO_ROOT = Path(__file__).resolve().parents[2]
_SCHEMA_ID = "graphrag.fmea.v1"


@dataclass(frozen=True, slots=True)
class DomainPackSpec:
    manifest_path: Path
    template_path: Path
    scoring_path: Path
    extension_values: tuple[tuple[str, str, object], ...]


PACK_SPECS: dict[str, DomainPackSpec] = {
    "fuel-combustion": DomainPackSpec(
        manifest_path=REPO_ROOT / "domain_packs" / "fuel-combustion" / "manifest.yaml",
        template_path=REPO_ROOT / "templates" / "examples" / "fuel-combustion-fmea.yaml",
        scoring_path=REPO_ROOT / "domain_packs" / "fuel-combustion" / "scoring" / "sod-rpn-1.0.0.yaml",
        extension_values=(),
    ),
    "electrical-demo": DomainPackSpec(
        manifest_path=REPO_ROOT / "domain_packs" / "electrical-demo" / "manifest.yaml",
        template_path=REPO_ROOT / "domain_packs" / "electrical-demo" / "templates" / "fmea.yaml",
        scoring_path=REPO_ROOT / "domain_packs" / "electrical-demo" / "scoring" / "sod-rpn.yaml",
        extension_values=(
            ("electrical.voltage", "decimal", "230.0"),
            ("electrical.current", "decimal", "12.5"),
            ("electrical.isolation", "string", "reinforced"),
        ),
    ),
    "software-demo": DomainPackSpec(
        manifest_path=REPO_ROOT / "domain_packs" / "software-demo" / "manifest.yaml",
        template_path=REPO_ROOT / "domain_packs" / "software-demo" / "templates" / "fmea.yaml",
        scoring_path=REPO_ROOT / "domain_packs" / "software-demo" / "scoring" / "sod-rpn.yaml",
        extension_values=(
            ("software.function", "string", "authorize_payment"),
            ("software.hazardous_behavior", "string", "retries_without_idempotency"),
            ("software.trigger", "string", "network_timeout"),
            ("software.detection", "string", "duplicate_transaction_alert"),
        ),
    ),
}


@dataclass(frozen=True, slots=True)
class DomainFixture:
    pack_id: str
    manifest: DomainPackManifest
    template: object
    scoring_rule: object
    row: FmeaRow
    row_payload: dict[str, object]
    snapshot: NormalizedFmeaSnapshot
    exports: dict[str, bytes]
    export_views: dict[str, dict[str, object]]
    generic_core_imported_domain_modules: tuple[str, ...]


def _typed_value(value: object, value_type: str) -> object:
    if value_type == "boolean":
        return json.loads(str(value).lower())
    if value_type == "integer":
        return int(value)  # pragma: no cover - demo packs currently use decimal/text values.
    return value


def _row_for(pack_id: str) -> FmeaRow:
    spec = PACK_SPECS[pack_id]
    values = tuple(FieldValue(key, value_type, _typed_value(value, value_type)) for key, value_type, value in spec.extension_values)
    claims = tuple(
        FieldClaim(key, ClaimStatus.KNOWN, EvidenceSupportStatus.SUPPORTED, ("evidence-demo-1",))
        for key, _, _ in spec.extension_values
    )
    return FmeaRow(
        row_id=f"{pack_id}-row-1",
        analysis_id=f"{pack_id}-analysis",
        evidence_pack_id=f"{pack_id}-evidence",
        item_id=f"{pack_id}-component",
        function_id=f"{pack_id}-function",
        failure_mode="bounded demonstration failure",
        causes=("fixture cause",),
        mechanisms=("fixture mechanism",),
        effects=("fixture effect",),
        symptoms=("fixture symptom",),
        controls=("fixture control",),
        barriers=("fixture barrier",),
        actions=("fixture action",),
        risk_assessment=None,
        field_evidence=(("failure_mode", ("evidence-demo-1",)),),
        field_support=(("failure_mode", EvidenceSupportStatus.SUPPORTED),),
        claim_status=ClaimStatus.KNOWN,
        review_status=ReviewStatus.ACCEPTED,
        publication_status=PublicationStatus.UNPUBLISHED,
        extension_values=values,
        field_claims=claims,
    )


def _parse_value(value: object, value_type: str) -> object:
    if value_type == "str":
        return "" if value is None else str(value)
    if value_type == "null":
        return None
    if value_type == "bool":
        return json.loads(str(value).lower())
    return json.loads(str(value))


def _parse_typed_table(rows: Any) -> list[dict[str, object]]:
    iterator = iter(tuple(tuple(row) for row in rows))
    headers = [str(value) for value in next(iterator)]
    type_column = headers.index("__types__")
    parsed: list[dict[str, object]] = []
    for row in iterator:
        if not any(value is not None and value != "" for value in row):
            continue
        types = json.loads(str(row[type_column]))
        item: dict[str, object] = {}
        for header, value in zip(headers, row, strict=True):
            if header in {"Identity", "__types__"} or header not in types:
                continue
            item[header] = _parse_value(value, types[header])
        parsed.append(item)
    return parsed


def _parse_xlsx(payload: bytes) -> dict[str, object]:
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    manifest_rows = list(workbook["Manifest"].iter_rows(values_only=True))
    manifest = {
        str(row[0]): _parse_value(row[1], str(row[2])) for row in manifest_rows[1:] if row[0] is not None
    }
    return {"rows": _parse_typed_table(workbook["FMEA"].iter_rows(values_only=True)), **manifest}


def _parse_docx(payload: bytes) -> dict[str, object]:
    document = Document(io.BytesIO(payload))
    marker_to_table = {}
    marker = None
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            text = Paragraph(child, document).text.strip()
            if text.startswith("Canonical table: "):
                marker = text.removeprefix("Canonical table: ")
        elif child.tag.endswith("}tbl") and marker is not None:
            marker_to_table[marker] = Table(child, document)
            marker = None
    # The marker is authoritative; the header check only avoids a misleading
    # KeyError when an old exporter omitted a marker entirely.
    assert {"Manifest", "FMEA"} <= set(marker_to_table)
    manifest_table = marker_to_table["Manifest"]
    manifest = {
        row.cells[0].text: _parse_value(row.cells[1].text, row.cells[2].text)
        for row in manifest_table.rows[1:]
    }
    return {
        "rows": _parse_typed_table(tuple(tuple(cell.text for cell in row.cells) for row in marker_to_table["FMEA"].rows)),
        **manifest,
    }


def _mapping_tuple(value: object) -> tuple[dict[str, object], ...]:
    return tuple(dict(item) for item in value)  # type: ignore[arg-type]


def _core_domain_imports_domain_modules() -> tuple[str, ...]:
    imported: set[str] = set()
    for path in (REPO_ROOT / "core_domain").rglob("*.py"):
        tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                imported.update(alias.name for alias in node.names if alias.name.startswith("domain_packs"))
            elif isinstance(node, ast.ImportFrom) and (node.module or "").startswith("domain_packs"):
                imported.add(node.module or "")
    return tuple(sorted(imported))


def run_domain_fixture(pack_id: str) -> DomainFixture:
    """Load, register, serialize, snapshot, and export one real pack fixture."""

    if pack_id not in PACK_SPECS:
        raise KeyError(pack_id)
    spec = PACK_SPECS[pack_id]
    manifest_source = spec.manifest_path.read_bytes()
    scoring_source = spec.scoring_path.read_bytes()
    manifest = load_domain_pack_manifest(manifest_source)
    scoring_rule = load_scoring_rule_pack(scoring_source)
    compiler = TemplateCompiler(schema_validator=Draft202012SchemaAdapter(), source_loader=load_template_source)
    template = compiler.compile_path(spec.template_path)

    with tempfile.TemporaryDirectory(prefix=f"fmea-{pack_id}-") as temp_dir:
        registry_root = Path(temp_dir)
        domain_registry = FileDomainPackRegistry(registry_root / "domain")
        scoring_registry = FileScoringRuleRegistry(registry_root / "scoring")
        template_registry = FileTemplateRegistry(registry_root / "template")
        domain_registry.register(manifest, manifest_source)
        scoring_registry.register(scoring_rule, scoring_source)
        template_registry.register(template, spec.template_path.read_bytes(), spec.template_path.suffix)
        manifest = domain_registry.get(manifest.pack_id, manifest.version)
        scoring_rule = scoring_registry.get(scoring_rule.rule_pack_id, scoring_rule.version)
        template = template_registry.get(template.metadata.template_id, template.metadata.version)

    row = _row_for(pack_id)
    if spec.extension_values:
        validate_extension_values(row, template)
    row_payload = json.loads(encode_json(row))
    revision = make_fmea_revision(
        analysis_id=f"{pack_id}-analysis",
        workspace_id="ws-1",
    )
    body_row = dict(row_payload)
    body_row["field_evidence"] = [
        {"field_key": key, "evidence_ids": evidence_ids}
        for key, evidence_ids in row_payload["field_evidence"]
    ]
    body_row["field_support"] = [
        {"field_key": key, "support_status": support_status}
        for key, support_status in row_payload["field_support"]
    ]
    body_row.setdefault("field_claims", [])
    body_row.setdefault("extension_values", [])
    body_row["row_hash"] = "a" * 64
    snapshot = make_normalized_snapshot(
        revision=revision,
        rows=(body_row,),
        risk_records=(),
        propagation=None,
        evidence_summary=({
            "pack_id": f"{pack_id}-evidence",
            "pack_hash": "b" * 64,
            "evidence_pack_version": "fixture-evidence-v1",
            "refs": [{
                "evidence_id": "evidence-demo-1",
                "document_id": f"{pack_id}-document",
                "document_version": "fixture-document-v1",
                "content_hash": "c" * 64,
                "evidence_hash": "d" * 64,
                "locator": {"page": 1, "span": 1},
                "quote": "Bounded structural demonstration evidence.",
                "source_type": "primary_document",
                "source_trust": "reviewed",
            }],
        },),
        decision_summary=({
            "record_type": "row_review",
            "decision_id": f"{pack_id}-decision-1",
            "workspace_id": "ws-1",
            "analysis_id": f"{pack_id}-analysis",
            "row_id": f"{pack_id}-row-1",
            "record_version": 1,
            "row_hash": "a" * 64,
            "role_category": "human_reviewer",
            "decision": "accepted",
            "reason": "Structural fixture accepted.",
            "decided_at": "2026-08-30T00:00:00Z",
        },),
        version_manifest={
            "schema_id": _SCHEMA_ID,
            "body_schema_version": "graphrag.fmea.body.v1",
            "domain_pack": {"id": manifest.pack_id, "version": manifest.version, "hash": manifest.content_hash},
            "template": {
                "id": template.metadata.template_id,
                "version": template.metadata.version,
                "hash": template.template_hash,
            },
            "scoring_rule": {
                "id": scoring_rule.rule_pack_id,
                "version": scoring_rule.version,
            },
        },
        audit_summary={"event_count": 1},
    )
    exports = {
        "json": CanonicalJsonExporter().render(snapshot),
        "xlsx": XlsxFmeaExporter().render(snapshot),
        "docx": DocxFmeaExporter().render(snapshot),
    }
    export_views = {
        "json": orjson.loads(exports["json"]),
        "xlsx": _parse_xlsx(exports["xlsx"]),
        "docx": _parse_docx(exports["docx"]),
    }
    return DomainFixture(
        pack_id=pack_id,
        manifest=manifest,
        template=template,
        scoring_rule=scoring_rule,
        row=row,
        row_payload=row_payload,
        snapshot=snapshot,
        exports=exports,
        export_views=export_views,
        generic_core_imported_domain_modules=_core_domain_imports_domain_modules(),
    )


@pytest.mark.parametrize("pack_id", ["electrical-demo", "software-demo"])
def test_demo_pack_sources_are_bundled_before_real_load(pack_id: str) -> None:
    spec = PACK_SPECS[pack_id]
    assert spec.manifest_path.is_file()
    assert spec.template_path.is_file()
    assert spec.scoring_path.is_file()


@pytest.mark.parametrize("pack_id", ["fuel-combustion", "electrical-demo", "software-demo"])
def test_domain_pack_uses_same_kernel_without_domain_imports(pack_id: str) -> None:
    result = run_domain_fixture(pack_id)
    assert result.manifest.compatible_schema_ids == (_SCHEMA_ID,)
    assert result.generic_core_imported_domain_modules == ()


@pytest.mark.parametrize("pack_id", ["electrical-demo", "software-demo"])
def test_demo_pack_loads_registers_and_compiles_authorized_artifacts(pack_id: str) -> None:
    result = run_domain_fixture(pack_id)

    assert result.manifest.pack_id == pack_id
    assert (result.template.metadata.template_id, result.template.metadata.version) in result.manifest.template_identities
    assert (result.scoring_rule.rule_pack_id, result.scoring_rule.version) in result.manifest.scoring_rule_identities
    assert result.manifest.extension_fields == tuple((key, value_type) for key, value_type, _ in PACK_SPECS[pack_id].extension_values)


def test_demo_packs_are_structurally_distinct_and_use_distinct_scoring_anchors() -> None:
    electrical = run_domain_fixture("electrical-demo")
    software = run_domain_fixture("software-demo")

    assert electrical.manifest.extension_fields != software.manifest.extension_fields
    assert electrical.template.output_schema != software.template.output_schema
    assert electrical.scoring_rule.rule_pack_id != software.scoring_rule.rule_pack_id
    assert electrical.scoring_rule.dimension_anchors != software.scoring_rule.dimension_anchors


@pytest.mark.parametrize("pack_id", ["fuel-combustion", "electrical-demo", "software-demo"])
def test_demo_exports_exercise_marked_publication_body_path(pack_id: str) -> None:
    result = run_domain_fixture(pack_id)

    assert result.snapshot.version_manifest["body_schema_version"] == "graphrag.fmea.body.v1"
    assert result.snapshot.decision_summary[0]["role_category"] == "human_reviewer"
    assert all(view["version_manifest"]["body_schema_version"] == "graphrag.fmea.body.v1" for view in result.export_views.values())


@pytest.mark.parametrize("pack_id", ["electrical-demo", "software-demo"])
def test_extension_values_and_claims_survive_row_snapshot_and_all_exports(pack_id: str) -> None:
    result = run_domain_fixture(pack_id)
    expected_values = {key: value for key, _, value in PACK_SPECS[pack_id].extension_values}
    expected_keys = tuple(expected_values)

    assert decode_row(encode_json(result.row)) == result.row
    assert tuple(item["field_key"] for item in result.row_payload["extension_values"]) == expected_keys
    assert tuple(item["field_key"] for item in result.row_payload["field_claims"]) == expected_keys
    assert {
        item["field_key"]: item["value"] for item in result.row_payload["extension_values"]
    } == expected_values
    assert all(
        item["claim_status"] == "known"
        and item["support_status"] == "supported"
        and item["evidence_ids"] == ["evidence-demo-1"]
        for item in result.row_payload["field_claims"]
    )
    assert _mapping_tuple(result.snapshot.rows[0]["extension_values"]) == _mapping_tuple(
        result.row_payload["extension_values"]
    )
    for view in result.export_views.values():
        assert _mapping_tuple(view["rows"][0]["extension_values"]) == _mapping_tuple(
            result.row_payload["extension_values"]
        )
        assert _mapping_tuple(view["rows"][0]["field_claims"]) == _mapping_tuple(result.row_payload["field_claims"])


def test_software_pack_has_no_physical_unit_propagation_surface() -> None:
    result = run_domain_fixture("software-demo")
    source = "\n".join(
        path.read_text(encoding="utf-8")
        for path in PACK_SPECS["software-demo"].manifest_path.parent.rglob("*")
        if path.is_file()
    )

    assert result.manifest.propagation_rule_identities == ()
    assert not any(token in source.casefold() for token in ("electrical.voltage", "electrical.current", "physical_unit"))
    assert "unit" not in result.template.output_schema.get("properties", {})
